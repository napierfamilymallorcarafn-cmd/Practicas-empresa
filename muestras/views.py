from django.http import HttpResponse, FileResponse, JsonResponse, StreamingHttpResponse
from .models import Muestra, Localizacion, Estudio, Envio, Documento, historial_estudios, historial_localizaciones,agenda_envio, registro_destruido, Congelador, Estante, Rack,Caja, Subposicion
from django.template import loader
from .forms import MuestraForm, UploadExcel, DocumentoForm, EstudioForm, Centroform, Congeladorform
from django.db import connection
from django.db import transaction
from django.contrib import messages  
from django.shortcuts import render,redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.safestring import mark_safe
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import json
import pandas as pd
import io,base64
# No quitar en caso de necesitar exportar las muestras en formato PDF
from reportlab.pdfgen import canvas
from django.conf import settings
from django.contrib.auth.models import User
import openpyxl,os
from django.db.models import Q
from django.utils import timezone 
from django.db.models import Count, Q, Prefetch
from django.core.exceptions import ObjectDoesNotExist
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
from datetime import date
from .parameters_config import get_upload_messages, get_excel_colors
import math

# Archivo que define las vistas de la aplicación, es decir, la manera de organizar,recoger y enviar al navegador los datos de los modelos del modelo 
# Los @ son decoradores de permisos, que limitan el acceso a las vistas de los distintos usuarios en base a si tienen ese permiso activado 

# ============================================================================
# FUNCIONES AUXILIARES PARA PANTALLA DE PROGRESO (StreamingHttpResponse)
# ============================================================================

def _progress_page_start(titulo, total):
    """
    Devuelve el HTML inicial de la página de progreso como popup/modal.
    Se envía al navegador como primer bloque del streaming.
    """
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{titulo}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', Arial, sans-serif; background: transparent; min-height: 100vh; }}
        .overlay {{ position: fixed; inset: 0; background: rgba(0,0,0,0.55); backdrop-filter: blur(4px); -webkit-backdrop-filter: blur(4px); display: flex; justify-content: center; align-items: center; z-index: 99999; animation: fadeIn 0.3s ease; }}
        @keyframes fadeIn {{ from {{ opacity: 0; }} to {{ opacity: 1; }} }}
        @keyframes slideUp {{ from {{ opacity: 0; transform: translateY(30px) scale(0.97); }} to {{ opacity: 1; transform: translateY(0) scale(1); }} }}
        .popup {{ background: white; border-radius: 14px; box-shadow: 0 20px 60px rgba(0,0,0,0.3), 0 0 0 1px rgba(255,255,255,0.1); padding: 36px 34px 30px; text-align: center; max-width: 440px; width: 90%; animation: slideUp 0.35s ease; position: relative; }}
        .popup h2 {{ color: #1a1a2e; margin-bottom: 22px; font-size: 18px; font-weight: 600; }}
        .bar-wrapper {{ background: #e9ecef; border-radius: 8px; height: 26px; overflow: hidden; margin-bottom: 10px; position: relative; }}
        .bar {{ background: linear-gradient(90deg, #4361ee, #3a0ca3); height: 100%; border-radius: 8px; transition: width 0.25s ease; width: 0%; display: flex; align-items: center; justify-content: center; color: white; font-size: 12px; font-weight: 700; min-width: 40px; }}
        .bar.done {{ background: linear-gradient(90deg, #2dc653, #1b9e3e); }}
        .bar.error {{ background: linear-gradient(90deg, #e63946, #c1121f); }}
        .info {{ color: #555; font-size: 14px; margin-top: 10px; min-height: 22px; }}
        .spinner {{ display: inline-block; width: 16px; height: 16px; border: 2.5px solid #ddd; border-top-color: #4361ee; border-radius: 50%; animation: spin 0.7s linear infinite; vertical-align: middle; margin-right: 6px; }}
        @keyframes spin {{ to {{ transform: rotate(360deg); }} }}
        .counter {{ color: #888; font-size: 12px; margin-top: 5px; }}
        .icon-ok {{ display: none; font-size: 34px; margin-bottom: 8px; }}
        .icon-ok.show {{ display: block; animation: slideUp 0.3s ease; }}
    </style>
</head>
<body>
    <div class="overlay">
        <div class="popup">
            <div class="icon-ok" id="picon"></div>
            <h2>{titulo}</h2>
            <div class="bar-wrapper">
                <div class="bar" id="pbar">0%</div>
            </div>
            <p class="info" id="ptxt"><span class="spinner"></span>Preparando {total} registros&hellip;</p>
            <p class="counter" id="pcnt">0 / {total}</p>
        </div>
    </div>
{'<!-- ' + 'x' * 1024 + ' -->'}
"""


def _progress_update(current, total):
    """
    Devuelve un bloque <script> que actualiza la barra de progreso.
    Enviar como bloque intermedio del streaming.
    """
    pct = (current / total * 100) if total > 0 else 100
    return f"""<script>
document.getElementById('pbar').style.width='{pct:.1f}%';
document.getElementById('pbar').textContent='{pct:.0f}%';
document.getElementById('ptxt').innerHTML='<span class="spinner"></span>Procesando fila {current} de {total}';
document.getElementById('pcnt').textContent='{current} / {total}';
</script>
"""


def _progress_done(redirect_url, message='¡Proceso completado correctamente!'):
    """
    Devuelve un bloque <script> que muestra finalización y redirige.
    Enviar como último bloque del streaming.
    """
    return f"""<script>
document.getElementById('pbar').style.width='100%';
document.getElementById('pbar').textContent='100%';
document.getElementById('pbar').classList.add('done');
document.getElementById('picon').textContent='\\u2705';
document.getElementById('picon').classList.add('show');
document.getElementById('ptxt').innerHTML='{message}';
document.getElementById('pcnt').textContent='Redirigiendo...';
setTimeout(function(){{
    document.querySelector('.overlay').style.transition='opacity 0.4s ease';
    document.querySelector('.overlay').style.opacity='0';
    setTimeout(function(){{ window.location.href='{redirect_url}'; }}, 400);
}}, 900);
</script>
</body></html>"""


def _progress_error(error_message, redirect_url=None):
    """
    Devuelve un bloque <script> que muestra error en el popup.
    """
    redir = ""
    if redirect_url:
        redir = f"""setTimeout(function(){{
    document.querySelector('.overlay').style.transition='opacity 0.4s ease';
    document.querySelector('.overlay').style.opacity='0';
    setTimeout(function(){{ window.location.href='{redirect_url}'; }}, 400);
}}, 2500);"""
    safe_msg = str(error_message).replace("'", "\\'").replace("\n", " ")
    return f"""<script>
document.getElementById('pbar').style.width='100%';
document.getElementById('pbar').textContent='Error';
document.getElementById('pbar').classList.add('error');
document.getElementById('picon').textContent='\\u274C';
document.getElementById('picon').classList.add('show');
document.getElementById('ptxt').innerHTML='Error: {safe_msg}';
document.getElementById('pcnt').textContent='';
{redir}
</script>
</body></html>"""


def _should_update(i, total, min_updates=80):
    """Devuelve True si debemos enviar actualización de progreso (evita enviar demasiados)."""
    if total <= min_updates:
        return True
    step = max(1, total // min_updates)
    return (i + 1) % step == 0 or i + 1 == total
@login_required
def principal(request):
    # Vista principal de la aplicación, muestra una página de bienvenida
    template = loader.get_template('principal.html')
    return HttpResponse(template.render(request=request))

# Vistas AJAX para cargar dinámicamente las localizaciones
@login_required
def get_estantes_por_congelador(request):
    """Devuelve los estantes disponibles para un congelador específico"""
    congelador_nombre = request.GET.get('congelador', '').strip()
    if not congelador_nombre:
        return JsonResponse({'error': 'Congelador no especificado'}, status=400)
    
    estantes = Estante.objects.filter(
        congelador__congelador__iexact=congelador_nombre
    ).values('id', 'numero').order_by('numero')
    
    return JsonResponse({'estantes': list(estantes)})

@login_required
def get_racks_por_estante(request):
    """Devuelve los racks disponibles para un estante específico"""
    estante_id = request.GET.get('estante_id')
    if not estante_id:
        return JsonResponse({'error': 'Estante no especificado'}, status=400)
    
    racks = Rack.objects.filter(
        estante_id=estante_id
    ).values('id', 'numero').order_by('numero')
    
    return JsonResponse({'racks': list(racks)})

@login_required
def get_cajas_por_rack(request):
    """Devuelve las cajas disponibles para un rack específico"""
    rack_id = request.GET.get('rack_id')
    if not rack_id:
        return JsonResponse({'error': 'Rack no especificado'}, status=400)
    
    cajas = Caja.objects.filter(
        rack_id=rack_id
    ).values('id', 'numero').order_by('numero')
    
    return JsonResponse({'cajas': list(cajas)})

@login_required
def get_subposiciones_por_caja(request):
    """Devuelve las subposiciones vacías disponibles para una caja específica"""
    caja_id = request.GET.get('caja_id')
    if not caja_id:
        return JsonResponse({'error': 'Caja no especificada'}, status=400)
    
    subposiciones = Subposicion.objects.filter(
        caja_id=caja_id,
        vacia=True
    ).values('id', 'numero').order_by('numero')
    
    return JsonResponse({'subposiciones': list(subposiciones)})

@login_required
def get_subposiciones_por_caja_tree(request):
    """Devuelve TODAS las subposiciones de una caja con info de muestra (para el árbol de archivo)."""
    caja_id = request.GET.get('caja_id')
    if not caja_id:
        return JsonResponse({'error': 'Caja no especificada'}, status=400)
    subposiciones = Subposicion.objects.filter(caja_id=caja_id).select_related('muestra').order_by('numero')
    resultado = []
    for s in subposiciones:
        item = {
            'id': s.id,
            'numero': s.numero,
            'vacia': s.vacia,
        }
        if not s.vacia and s.muestra:
            item['muestra_nom_lab'] = s.muestra.nom_lab
            item['muestra_estado'] = s.muestra.estado_actual or ''
        resultado.append(item)
    return JsonResponse({'subposiciones': resultado})

@login_required

# Vistas para Muestras
@permission_required('muestras.can_view_muestras_web')
def muestras_todas(request):
    # Vista que muestra todas las muestras y su localización asociada, requiere que el usuario esté autenticado
    muestras = (Muestra.objects.select_related('subposicion__caja__rack__estante__congelador'))

    # Filtrado de muestras si se proporcionan parámetros de búsqueda en los filtros del template
    field_names = [f.name for f in Muestra._meta.local_fields if f.name not in ('id','estudio')]
    field_names_readable = ['Id del individuo','Nombre dado por el laboratorio','Material','Volumen actual','Unidad de volumen','Concentración actual','Unidad de concentración','Masa actual','Unidad de masa','Fecha de extracción','Fecha de llegada','Observaciones','Estado inicial','Centro de procedencia','Lugar de procedencia','Estado actual']
    field_names_readable_dict = {k:v for (k,v) in zip(field_names,field_names_readable)}
    
    # Filtrado de campos de texto normales
    for field in field_names:
        val = request.GET.get(field)
        if val:
            # Separar por punto y coma (;) para permitir múltiples filtros
            valores = [v.strip() for v in val.split(';') if v.strip()]
            if valores:
                # Para campos que son dropdowns, usar búsqueda exacta
                if field in ['id_material', 'centro_procedencia', 'lugar_procedencia', 'estado_actual']:
                    q_filters = Q()
                    for valor in valores:
                        if valor == 'null':
                            # Incluir tanto NULL como cadenas vacías
                            q_filters |= Q(**{f"{field}__isnull": True}) | Q(**{f"{field}": ""})
                        else:
                            q_filters |= Q(**{f"{field}": valor})
                    muestras = muestras.filter(q_filters)
                else:
                    # Para campos de texto: búsqueda icontains para cada valor
                    q_filters = Q()
                    for valor in valores:
                        q_filters |= Q(**{f"{field}__icontains": valor})
                    muestras = muestras.filter(q_filters)
    
    # Filtrado de las muestras en base al estudio específico (por ID)
    if request.GET.get('estudio'):
        estudio_val = request.GET['estudio']
        if estudio_val:
            # Separar por punto y coma para múltiples estudios
            estudios_ids = [e.strip() for e in estudio_val.split(';') if e.strip()]
            if estudios_ids:
                q_filters = Q()
                for estudio_id in estudios_ids:
                    if estudio_id == 'null':
                        q_filters |= Q(estudio__isnull=True)
                    else:
                        try:
                            q_filters |= Q(estudio__id=int(estudio_id))
                        except (ValueError, TypeError):
                            pass
                muestras = muestras.filter(q_filters)

    # Filtrado de material por texto libre (icontains, separado por ;)
    if request.GET.get('id_material_texto'):
        valores = [v.strip() for v in request.GET['id_material_texto'].split(';') if v.strip()]
        if valores:
            q_filters = Q()
            for valor in valores:
                q_filters |= Q(id_material__icontains=valor)
            muestras = muestras.filter(q_filters)

    # Filtrado de estudio por texto libre (nombre del estudio, icontains, separado por ;)
    if request.GET.get('estudio_texto'):
        valores = [v.strip() for v in request.GET['estudio_texto'].split(';') if v.strip()]
        if valores:
            q_filters = Q()
            for valor in valores:
                if valor.lower() == 'null' or valor.lower() == 'sin estudio':
                    q_filters |= Q(estudio__isnull=True)
                else:
                    q_filters |= Q(estudio__nombre_estudio__icontains=valor)
            muestras = muestras.filter(q_filters)

    # Filtrado de centro de procedencia por texto libre (icontains, separado por ;)
    if request.GET.get('centro_procedencia_texto'):
        valores = [v.strip() for v in request.GET['centro_procedencia_texto'].split(';') if v.strip()]
        if valores:
            q_filters = Q()
            for valor in valores:
                q_filters |= Q(centro_procedencia__icontains=valor)
            muestras = muestras.filter(q_filters)

    # Filtrado de lugar de procedencia por texto libre (icontains, separado por ;)
    if request.GET.get('lugar_procedencia_texto'):
        valores = [v.strip() for v in request.GET['lugar_procedencia_texto'].split(';') if v.strip()]
        if valores:
            q_filters = Q()
            for valor in valores:
                q_filters |= Q(lugar_procedencia__icontains=valor)
            muestras = muestras.filter(q_filters)

    # Filtrado de estado actual por texto libre (icontains, separado por ;)
    if request.GET.get('estado_actual_texto'):
        valores = [v.strip() for v in request.GET['estado_actual_texto'].split(';') if v.strip()]
        if valores:
            q_filters = Q()
            for valor in valores:
                q_filters |= Q(estado_actual__icontains=valor)
            muestras = muestras.filter(q_filters)
    
    # Filtrado por localizaciones
    if request.GET.get('congelador'):
        congelador_val = request.GET['congelador']
        congeladores = [c.strip() for c in congelador_val.split(';') if c.strip()]
        if congeladores:
            q_filters = Q()
            for congelador in congeladores:
                if congelador == 'null':
                    q_filters |= Q(subposicion__isnull=True)
                else:
                    q_filters |= Q(subposicion__caja__rack__estante__congelador__congelador__iexact=congelador)
            muestras = muestras.filter(q_filters)
    
    if request.GET.get('estante'):
        estante_val = request.GET['estante']
        estantes = [e.strip() for e in estante_val.split(';') if e.strip()]
        if estantes:
            q_filters = Q()
            for estante in estantes:
                if estante == 'null':
                    q_filters |= Q(subposicion__isnull=True)
                else:
                    try:
                        q_filters |= Q(subposicion__caja__rack__estante__numero=int(estante))
                    except (ValueError, TypeError):
                        pass
            muestras = muestras.filter(q_filters)
    
    if request.GET.get('rack'):
        rack_val = request.GET['rack']
        racks = [r.strip() for r in rack_val.split(';') if r.strip()]
        if racks:
            q_filters = Q()
            for rack in racks:
                if rack == 'null':
                    q_filters |= Q(subposicion__isnull=True)
                else:
                    q_filters |= Q(subposicion__caja__rack__numero__iexact=rack)
            muestras = muestras.filter(q_filters)
    
    if request.GET.get('caja'):
        caja_val = request.GET['caja']
        cajas = [c.strip() for c in caja_val.split(';') if c.strip()]
        if cajas:
            q_filters = Q()
            for caja in cajas:
                if caja == 'null':
                    q_filters |= Q(subposicion__isnull=True)
                else:
                    q_filters |= Q(subposicion__caja__numero__iexact=caja)
            muestras = muestras.filter(q_filters)

    # Filtrado de las muestras a mostrar si el perfil es de un investigador, mostrando solo las asociadas a sus estudios
    if request.user.groups.filter(name='Investigadores'):
        muestras = Muestra.objects.filter(Q(estudio__investigadores_asociados__username=request.user))
    
    # Búsqueda general (server-side): filtra en todos los campos de texto relevantes
    busqueda_general = request.GET.get('busqueda', '').strip()
    if busqueda_general:
        q_busqueda = Q()
        q_busqueda |= Q(id_individuo__icontains=busqueda_general)
        q_busqueda |= Q(nom_lab__icontains=busqueda_general)
        q_busqueda |= Q(id_material__icontains=busqueda_general)
        q_busqueda |= Q(observaciones__icontains=busqueda_general)
        q_busqueda |= Q(estado_inicial__icontains=busqueda_general)
        q_busqueda |= Q(centro_procedencia__icontains=busqueda_general)
        q_busqueda |= Q(lugar_procedencia__icontains=busqueda_general)
        q_busqueda |= Q(estado_actual__icontains=busqueda_general)
        q_busqueda |= Q(estudio__nombre_estudio__icontains=busqueda_general)
        muestras = muestras.filter(q_busqueda)

    '''
    # Crear un PDF con las muestras filtradas
    if request.GET.get('crear_pdf'):    
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        y = 800
        p.setFont("Helvetica", 16)
        p.drawString(30,y, "Listado de Muestras")
        p.setFont("Helvetica", 12)
        y -= 30
        p.drawString(30, y, "ID Individuo")
        p.drawString(150, y, "Nombre Laboratorio")
        p.drawString(300, y, "Localización")
        y-= 30
        for muestra in muestras:
            p.drawString(30, y, muestra.id_individuo)
            p.drawString(150, y, muestra.nom_lab)
            p.drawString(300, y, str(muestra.localizacion.first()) if muestra.localizacion.exists() else 'No archivada')
            y -= 20
            if y < 50:
                p.showPage()
                y = 800
        p.save()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='listado_muestras.pdf')
    '''



     # Crear un Excel con las muestras filtradas 
    if request.GET.get('exportar_excel'):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="listado_muestras.xlsx"'
        wb = openpyxl.load_workbook(os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'listado_muestras.xlsx'))
        ws = wb.active
        row_num = 2
        for muestra in muestras:
            col_num = 1
            
            for field in field_names:
                value = muestra.__dict__[field]
                if value is None:
                    value = ''
                ws.cell(row_num, col_num).value= str(value)
                col_num += 1
            value = muestra.estudio.nombre_estudio if muestra.estudio else ''
            ws.cell(row_num, col_num).value= str(value)
            col_num += 1
            value = muestra.posicion_completa()
            if value is None:
                value = ''
            else:
                value = value.split("-")
                for columna in value:
                    ws.cell(row_num, col_num).value= str(columna)
                    col_num += 1
            row_num += 1
    
        wb.save(response)
        return response
    
    # Obtener opciones para los dropdowns
    opciones_materiales = Muestra.objects.values_list('id_material', flat=True).distinct().exclude(id_material__isnull=True).exclude(id_material='')
    opciones_centros = Muestra.objects.values_list('centro_procedencia', flat=True).distinct().exclude(centro_procedencia__isnull=True).exclude(centro_procedencia='')
    opciones_lugares = Muestra.objects.values_list('lugar_procedencia', flat=True).distinct().exclude(lugar_procedencia__isnull=True).exclude(lugar_procedencia='')
    opciones_estudios = Estudio.objects.all()
    opciones_congeladores = Congelador.objects.values_list('congelador', flat=True).distinct()
    opciones_estantes = Estante.objects.values_list('numero', flat=True).distinct().order_by('numero')
    opciones_racks = Rack.objects.values_list('numero', flat=True).distinct().order_by('numero')
    opciones_cajas = Caja.objects.values_list('numero', flat=True).distinct().order_by('numero')
    

    # Paginación
    contador_total = muestras.count()
    items_por_pagina = request.GET.get('items_por_pagina', 25)
    if str(items_por_pagina) == 'todas':
        items_por_pagina = 'todas'
        paginator = Paginator(muestras, max(contador_total, 1))
    else:
        try:
            items_por_pagina = int(items_por_pagina)
            if items_por_pagina not in [10, 25, 50, 100]:
                items_por_pagina = 25
        except (ValueError, TypeError):
            items_por_pagina = 25
        paginator = Paginator(muestras, items_por_pagina)
    numero_pagina = request.GET.get('page', 1)
    
    try:
        muestras_pagina = paginator.page(numero_pagina)
    except PageNotAnInteger:
        muestras_pagina = paginator.page(1)
    except EmptyPage:
        muestras_pagina = paginator.page(paginator.num_pages)

    # Cargar el template y pasar las muestras y los campos de filtro al mismo
    template = loader.get_template('muestras_todas.html')
    context = {    
        'muestras': muestras,
        'contador_muestras': muestras.count(),
        'field_names': field_names,
        'field_names_readable_dict': field_names_readable_dict,
        'opciones_materiales': opciones_materiales,
        'opciones_centros': opciones_centros,
        'opciones_lugares': opciones_lugares,
        'opciones_estudios': opciones_estudios,
        'opciones_congeladores': opciones_congeladores,
        'opciones_estantes': opciones_estantes,
        'opciones_racks': opciones_racks,
        'opciones_cajas': opciones_cajas,
        'muestras_page': muestras_pagina,
        'paginator': paginator,
        'items_por_pagina': items_por_pagina,
        'busqueda': busqueda_general,
        'muestras_ids_json': json.dumps(list(muestras.values_list('id', flat=True))),
    }
    return HttpResponse(template.render(context, request))
@login_required
@permission_required('muestras.can_view_muestras_web')
def acciones_post(request):
    # Vista que redirigue la petición del usuario según el botón de acción que haya pulsado y las muestras que haya seleccionado
    if request.method=="POST":
        muestras_seleccionadas = request.POST.getlist('muestra_id')
        if 'estudio' in request.POST:
            # Se guardan las muestras seleccionadas en la sesión y se redirigue al usuario a la selección de un estudio
            if muestras_seleccionadas:
                request.session['muestras_estudio']=muestras_seleccionadas
                return redirect('seleccionar_estudio')
        elif 'eliminar' in request.POST:
            # Se eliminan las muestras seleccionadas
            if muestras_seleccionadas:
                muestras_a_procesar = Muestra.objects.filter(id__in=muestras_seleccionadas)
                for muestra in muestras_a_procesar:
                    eliminar_muestra(request, muestra.nom_lab)
        elif 'envio' in request.POST:
            # Se guardan las muestras seleccionadas en la sesión y se redirigue al usuario a la agenda de envíos
            if 'muestras_envio' in request.session:
                del request.session['muestras_envio']
            for muestra in muestras_seleccionadas:
                # Se eliminan las muestras destruidas de la lista de envío
                if Muestra.objects.get(id=muestra).estado_actual == 'DEST':
                    muestras_seleccionadas.remove(muestra)
            request.session['muestras_envio']=muestras_seleccionadas
            return redirect('agenda')
        elif 'destruir' in request.POST:
            # Se marcan las muestras seleccionadas como destruidas
            if muestras_seleccionadas:
                numero_muestras_destruidas = 0
                muestras_a_destruir = Muestra.objects.filter(id__in=muestras_seleccionadas)
                for sample in muestras_a_destruir:
                    sample.estado_actual = 'DEST'
                    sample.volumen_actual = 0
                    sample.concentracion_actual = 0
                    sample.save()
                    # Liberar la subposición asociada
                    if Subposicion.objects.filter(muestra=sample).exists():
                        subposicion = Subposicion.objects.get(muestra=sample)
                        subposicion.vacia = True
                        subposicion.muestra = None
                        subposicion.save()
                    if Localizacion.objects.filter(muestra=sample).exists():
                        # Actualizar todas las localizaciones de esta muestra
                        Localizacion.objects.filter(muestra=sample).update(muestra=None)
                    registro_destruccion = registro_destruido.objects.create(muestra = sample,
                                                                             fecha = timezone.now(),
                                                                             usuario = request.user)
                    registro_destruccion.save()
                    numero_muestras_destruidas += 1
                messages.success(request, f'{numero_muestras_destruidas} muestras destruidas correctamente. ')
        elif 'cambio_posicion' in request.POST:
            # Se redirigue al usuario a la vista de cambio de posición de muestras
            return redirect('cambio_posicion')
        elif 'exportar_seleccionadas' in request.POST:
            # Exportar a Excel solo las muestras seleccionadas
            if muestras_seleccionadas:
                field_names = [f.name for f in Muestra._meta.local_fields if f.name not in ('id','estudio')]
                muestras_export = Muestra.objects.filter(id__in=muestras_seleccionadas)
                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename="muestras_seleccionadas.xlsx"'
                wb = openpyxl.load_workbook(os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'listado_muestras.xlsx'))
                ws = wb.active
                row_num = 2
                for muestra in muestras_export:
                    col_num = 1
                    for field in field_names:
                        value = muestra.__dict__[field]
                        if value is None:
                            value = ''
                        ws.cell(row_num, col_num).value = str(value)
                        col_num += 1
                    value = muestra.estudio.nombre_estudio if muestra.estudio else ''
                    ws.cell(row_num, col_num).value = str(value)
                    col_num += 1
                    value = muestra.posicion_completa()
                    if value is None:
                        value = ''
                    else:
                        value = value.split("-")
                        for columna in value:
                            ws.cell(row_num, col_num).value = str(columna)
                            col_num += 1
                    row_num += 1
                wb.save(response)
                return response             
    return redirect('muestras_todas')    
def detalles_muestra(request, nom_lab):
    # Vista que muestra los detalles de una muestra específica, requiere permiso para ver muestras
    muestra = Muestra.objects.get(nom_lab=nom_lab)
    template = loader.get_template('detalles_muestra.html')
    context = {
        'muestra': muestra,
    }
    return HttpResponse(template.render(context, request))
@permission_required('muestras.can_add_muestras_web')


@permission_required('muestras.can_add_muestras_web')
def añadir_muestras(request):
    # Vista para añadir una nueva muestra, requiere permiso para añadir muestras
    if request.method == 'POST':
        form_muestra = MuestraForm(request.POST)
        if form_muestra.is_valid():
            muestra = form_muestra.save()
            
            # Manejar localización si se proporcionó
            subposicion_id = request.POST.get("subposicion_id")
            if subposicion_id:
                try:
                    subposicion = Subposicion.objects.select_for_update().get(id=subposicion_id)
                    
                    if subposicion.vacia:
                        # Asignar la muestra a la subposición y crear la localización
                        subposicion.muestra = muestra
                        subposicion.vacia = False
                        subposicion.save()
                        
                        localizacion = Localizacion.objects.create(
                            muestra=muestra,
                            congelador=subposicion.caja.rack.estante.congelador.congelador,
                            estante=subposicion.caja.rack.estante.numero,
                            rack=subposicion.caja.rack.numero,
                            caja=subposicion.caja.numero,
                            subposicion=subposicion.numero,
                        )
                        
                        historial_localizaciones.objects.create(
                            muestra=muestra,
                            localizacion=localizacion,
                            fecha_asignacion=timezone.now(),
                            usuario_asignacion=request.user
                        )
                        messages.success(request, 'Muestra añadida correctamente')
                    else:
                        messages.error(request, 'La subposición está ocupada por otra muestra, la muestra se guardará sin localización.')
                except Subposicion.DoesNotExist:
                    messages.error(request, 'La subposición seleccionada no existe.')
            else:
                messages.success(request, 'Muestra añadida correctamente')
            
            return redirect('muestras_todas')
    else:
        form_muestra = MuestraForm()
    
    # Obtener congeladores para los menús desplegables
    congeladores = Congelador.objects.all().values('id', 'congelador').order_by('congelador')
    
    return render(request, 'añadir_muestras.html', {
        'form_muestra': form_muestra,
        'congeladores': congeladores,
    })


@permission_required('muestras.can_delete_muestras_web')
def eliminar_muestra(request, nom_lab):
    # Vista para eliminar una muestra, requiere permiso para eliminar muestras
    muestra = get_object_or_404(Muestra, nom_lab=nom_lab)
    if Subposicion.objects.filter(muestra = muestra).exists():
        # Liberar la subposición asociada a la muestra antes de eliminarla
        subposicion = Subposicion.objects.get(muestra = muestra)
        subposicion.muestra = None
        subposicion.vacia = True
        subposicion.save()
    muestra.delete()
    messages.success(request,'Muestras eliminadas correctamente')
    return redirect('muestras_todas')
@permission_required('muestras.can_add_muestras_web')
def upload_excel(request):
    # Mostrar confirmación después de validación con progreso
    if request.GET.get('mostrar_confirmacion') and 'confirmacion_pendiente' in request.session:
        datos_conf = request.session.pop('confirmacion_pendiente')
        for msg in datos_conf.get('mensajes', []):
            getattr(messages, msg['level'])(request, msg['text'])
        return render(request, datos_conf['template'], datos_conf.get('context', {}))
    # Vista para subir un archivo Excel con múltiples muestras y asociarlas a un estudio y subposición desde el excel, requiere permiso para añadir muestras
    if request.method=="POST":
        form = UploadExcel(request.POST, request.FILES)
        # Si el usuario confirma, se crean en la base de datos los registros validos
        if 'confirmar' in request.POST:
            filas_validas = request.session.get('filas_validas',[])
            errores_sesion = request.session.get('errores', {})
            total = len(filas_validas)

            def gen_confirmar_muestras():
                yield _progress_page_start('Importando muestras', total)
                try:
                    with transaction.atomic():
                        for i, datos in enumerate(filas_validas):
                            # Antes de crear, eliminar del dict los campos marcados como advertencia
                            fila_num = datos.get('fila')
                            advertencias = []
                            if fila_num:
                                advertencias = errores_sesion.get(fila_num, {}).get('advertencias', [])
                            for warn in advertencias:
                                if warn == 'fecha_incoherente':
                                    datos['fecha_extraccion'] = None
                                    datos['fecha_llegada'] = None
                                elif ':' in warn:
                                    _, campo = warn.split(':', 1)
                                    if campo in datos:
                                        datos[campo] = None

                            fecha_extraccion = None
                            fecha_llegada = None
                            if datos.get("fecha_extraccion"):
                                try:
                                    fecha_extraccion = date.fromisoformat(datos["fecha_extraccion"])
                                except Exception:
                                    fecha_extraccion = None
                            if datos.get("fecha_llegada"):
                                try:
                                    fecha_llegada = date.fromisoformat(datos["fecha_llegada"])
                                except Exception:
                                    fecha_llegada = None
                            estudio = None
                            if datos.get("estudio"):
                                estudio = Estudio.objects.get(id=datos["estudio"])

                            muestra = Muestra.objects.create(
                                id_individuo=datos.get("id_individuo"),
                                nom_lab=datos["nom_lab"],
                                id_material=datos.get("id_material"),
                                volumen_actual=datos.get("volumen_actual"),
                                unidad_volumen=datos.get("unidad_volumen"),
                                concentracion_actual=datos.get("concentracion_actual"),
                                unidad_concentracion=datos.get("unidad_concentracion"),
                                masa_actual=datos.get("masa_actual"),
                                unidad_masa=datos.get("unidad_masa"),
                                fecha_extraccion=fecha_extraccion,
                                fecha_llegada=fecha_llegada,
                                observaciones=datos.get("observaciones"),
                                estado_inicial=datos.get("estado_inicial"),
                                centro_procedencia=datos.get("centro_procedencia"),
                                lugar_procedencia=datos.get("lugar_procedencia"),
                                estado_actual=datos.get("estado_actual") or None,
                                estudio=estudio,
                            )

                            subposicion_id = datos.get("subposicion_id")
                            if subposicion_id:
                                try:
                                    subposicion = Subposicion.objects.select_for_update().get(id=subposicion_id)
                                except Subposicion.DoesNotExist:
                                    subposicion = None
                                if subposicion:
                                    localizacion = Localizacion.objects.create(
                                        muestra=muestra,
                                        congelador=subposicion.caja.rack.estante.congelador.congelador,
                                        estante=subposicion.caja.rack.estante.numero,
                                        rack=subposicion.caja.rack.numero,
                                        caja=subposicion.caja.numero,
                                        subposicion=subposicion.numero,
                                    )
                                    subposicion.vacia = False
                                    subposicion.muestra = muestra
                                    subposicion.save()
                                    historial_localizaciones.objects.create(
                                        muestra=muestra,
                                        localizacion=localizacion,
                                        fecha_asignacion=timezone.now(),
                                        usuario_asignacion=request.user
                                    )

                            if datos.get("estudio"):
                                historial_estudios.objects.create(
                                    muestra=muestra,
                                    estudio=estudio,
                                    fecha_asignacion=timezone.now(),
                                    usuario_asignacion=request.user
                                )
                            if _should_update(i, total):
                                yield _progress_update(i + 1, total)
                    yield _progress_done('/muestras/', 'Muestras importadas correctamente')
                except Exception as e:
                    yield _progress_error(str(e))

            return StreamingHttpResponse(gen_confirmar_muestras(), content_type='text/html')
        # Si el usuario cancela, no se añade nada a la base de datos
        elif 'cancelar' in request.POST:
            messages.error(request,'Las muestras no se han añadido')
            return redirect('muestras_todas')
        
        # Si hay un archivo excel subido, se procesa
        elif 'excel_file' in request.FILES:
            # Limpiar sesión residual de uploads anteriores
            if 'columnas_adicionales' in request.session:
                del request.session['columnas_adicionales']
            if form.is_valid():
                # Leer excel y preparar columnas 
                excel_file = request.FILES['excel_file']
                # Validar que sea un archivo Excel
                if not excel_file.name.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                    return render(request, 'upload_excel.html', {'form': form, 'error': '❌ Error de formato: El archivo debe ser un Excel (.xlsx, .xls o .xlsm).'})
                
                excel_bytes = excel_file.read()
                request.session['excel_file_name'] = excel_file.name
                request.session['excel_file_base64']= base64.b64encode(excel_bytes).decode()
                excel_stream = io.BytesIO(excel_bytes)
                
                # Intentar leer el archivo Excel
                try:
                    df = pd.read_excel(excel_stream)
                except Exception as e:
                    return render(request, 'upload_excel.html', {'form': form, 'error': f'❌ Error al leer el archivo Excel: {str(e)}'})
                
                # Validar que tenga al menos una fila de datos
                if df.empty or len(df) == 0:
                    return render(request, 'upload_excel.html', {'form': form, 'error': '❌ Error de formato: El archivo Excel está vacío o no contiene filas de datos.'})
                
                rename_columns = {
                    'ID Individuo': 'id_individuo', 
                    'Nombre Laboratorio': 'nom_lab',
                    'ID Material': 'id_material',
                    'Volumen Actual': 'volumen_actual',
                    'Unidad Volumen': 'unidad_volumen',
                    'Concentracion Actual': 'concentracion_actual',
                    'Unidad Concentracion': 'unidad_concentracion',
                    'Masa Actual': 'masa_actual',
                    'Unidad Masa': 'unidad_masa',
                    'Fecha Extraccion': 'fecha_extraccion',
                    'Fecha Llegada': 'fecha_llegada',
                    'Observaciones': 'observaciones',
                    'Estado Inicial': 'estado_inicial',
                    'Centro Procedencia': 'centro_procedencia',
                    'Lugar Procedencia': 'lugar_procedencia',
                    'Estado actual': 'estado_actual',
                    'Congelador': 'congelador', 
                    'Estante': 'estante',
                    'Posición del rack en el estante': 'posicion_rack_estante',
                    'Rack': 'rack',
                    'Posición de la caja en el rack': 'posicion_caja_rack',
                    'Caja': 'caja',
                    'Subposición': 'subposicion',
                    'Estudio':'estudio'
                }
                # Validar columnas
                columnas_esperadas = set(rename_columns.keys())
                columnas_existentes = set(df.columns)
                columnas_faltantes = columnas_esperadas - columnas_existentes
                if columnas_faltantes:
                    columnas_str = ', '.join(sorted(columnas_faltantes))
                    return render(request, 'upload_excel.html', {'form': form, 'error': f'❌ Error de formato: El archivo Excel no contiene las siguientes columnas esperadas: {columnas_str}'})
                
                columnas_adicionales = columnas_existentes - columnas_esperadas
                if columnas_adicionales:
                    columnas_adicionales_str = ', '.join(sorted(columnas_adicionales))
                    request.session['columnas_adicionales'] = columnas_adicionales_str
                
                df.rename(columns=rename_columns, inplace=True)
                # Funciones para normalizar las columnas del excel
                def norm(value):
                    if value is None or pd.isna(value):
                        return None

                    if isinstance(value, str):
                        value = value.strip()
                        return value if value != "" else None

                    return value
                
                def norm_code(value):
                    if value is None or pd.isna(value):
                        return None

                    if isinstance(value, float) and value.is_integer():
                        return str(int(value))

                    return str(value).strip()

                # Definir campos obligatorios para una muestra (solo nombre de laboratorio)
                obligatorios = ["nom_lab"]

                cache = {
                    'subposiciones': {
                        (
                            str(c.caja.rack.estante.congelador.congelador).lower(),
                            c.caja.rack.estante.numero,
                            str(c.caja.rack.numero).lower(),
                            str(c.caja.numero).lower(),
                            str(c.numero).lower()
                        ): c
                         for c in Subposicion.objects.select_related('caja__rack__estante__congelador')
                    },

                    'estudios':{e:Estudio.objects.get(id=e)
                                for e in Estudio.objects.values_list('id', flat=True)},

                    'muestras_existentes': set(Muestra.objects.values_list('nom_lab',flat=True))
                }

                total_filas = len(df)

                def gen_validar_muestras():
                    errores = {}
                    filas_validas = []
                    numero_registros = 0
                    nom_lab_excel = set()

                    yield _progress_page_start('Validando muestras', total_filas)

                    # Recorrer el df para detectar errores y normalizar
                    for idx, row in df.iterrows():
                        numero_registros += 1
                        fila = idx + 2 
                        errores[fila]={"bloqueantes":[],"advertencias":[]}
                        datos = {
                            "id_individuo":norm(row['id_individuo']),
                            "nom_lab":norm(row['nom_lab']),
                            "id_material":norm(row['id_material']),
                            "volumen_actual":norm_code(row['volumen_actual']),
                            "unidad_volumen":norm(row['unidad_volumen']),
                            "concentracion_actual":norm_code(row['concentracion_actual']),
                            "unidad_concentracion":norm(row['unidad_concentracion']),
                            "masa_actual":norm_code(row['masa_actual']),
                            "unidad_masa":norm(row['unidad_masa']),
                            "fecha_extraccion":norm(row['fecha_extraccion']),
                            "fecha_llegada":norm(row['fecha_llegada']),
                            "observaciones":norm(row['observaciones']),
                            "estado_inicial":norm(row['estado_inicial']),
                            "centro_procedencia":norm(row['centro_procedencia']),
                            "lugar_procedencia":norm(row['lugar_procedencia']),
                            "estado_actual":norm(row['estado_actual']),
                            "congelador":(norm_code(row['congelador']) or None),
                            "estante":norm_code(row['estante']),
                            "posicion_rack_estante":norm_code(row['posicion_rack_estante']),
                            "rack":(norm_code(row['rack']) or None),
                            "posicion_caja_rack":norm_code(row['posicion_caja_rack']),
                            "caja":(norm_code(row['caja']) or None),
                            "subposicion":(norm_code(row['subposicion']) or None),
                            "estudio":norm_code(row['estudio'])  
                        }
                        # Normalizar a minúsculas los campos textuales relevantes
                        if datos['congelador'] is not None:
                            datos['congelador'] = str(datos['congelador']).lower()
                        if datos['rack'] is not None:
                            datos['rack'] = str(datos['rack']).lower()
                        if datos['caja'] is not None:
                            datos['caja'] = str(datos['caja']).lower()
                        if datos['subposicion'] is not None:
                            datos['subposicion'] = str(datos['subposicion']).lower()

                        for campo in obligatorios:
                            if not datos.get(campo):
                                errores[fila]["bloqueantes"].append(f"campo_obligatorio_vacio:{campo}")

                        # Validar que ningún campo contenga el carácter punto y coma (;)
                        campos_a_validar = ['id_individuo', 'nom_lab', 'id_material', 'unidad_volumen', 
                                           'unidad_concentracion', 'unidad_masa', 'observaciones', 'estado_inicial',
                                           'centro_procedencia', 'lugar_procedencia', 'estado_actual', 'congelador',
                                           'rack', 'caja', 'subposicion']
                        for campo in campos_a_validar:
                            valor = datos.get(campo)
                            if valor and isinstance(valor, str) and ';' in valor:
                                errores[fila]["bloqueantes"].append(f"caracter_invalido_semicolon:{campo}")

                        # Comprobar si los campos estan en el formato correcto (numéricos)
                        for campo in ['volumen_actual', 'concentracion_actual', 'masa_actual']:
                            if datos[campo] != None:
                                try:
                                    valor_numerico = float(datos[campo])
                                    # Validar que sea positivo o cero (≥0)
                                    if valor_numerico < 0:
                                        errores[fila]["advertencias"].append(f"valor_negativo:{campo}")
                                        datos[campo] = None  # No asignar valores negativos
                                    else:
                                        datos[campo] = valor_numerico
                                except (TypeError, ValueError):
                                    errores[fila]["advertencias"].append(f"formato_incorrecto:{campo}")
                                    datos[campo] = None  # No asignar valores con formato incorrecto
                        
                        # Validar estado_actual: debe ser uno de los valores permitidos (case-insensitive)
                        estado_actual_validos = {
                            'disponible': 'DISP',
                            'enviada': 'ENV',
                            'parcialmente enviada': 'PENV',
                            'enviada parcialmente': 'PENV',
                            'destruida': 'DEST',
                            'disp': 'DISP',
                            'env': 'ENV',
                            'penv': 'PENV',
                            'dest': 'DEST'
                        }
                        if datos.get('estado_actual'):
                            estado_normalizado = str(datos['estado_actual']).strip().lower()
                            if estado_normalizado in estado_actual_validos:
                                datos['estado_actual'] = estado_actual_validos[estado_normalizado]
                            else:
                                errores[fila]["advertencias"].append("estado_actual_invalido")
                                datos['estado_actual'] = None
                            
                        # Validar y normalizar fechas (esperado formato día-mes-año o fecha Excel)
                        for campo in ['fecha_extraccion', 'fecha_llegada']:
                            if datos[campo] != None:
                                try:
                                    # Forzar interpretación día/mes/año cuando sea una cadena
                                    fecha = pd.to_datetime(datos[campo], dayfirst=True, errors='raise')
                                    datos[campo] = fecha.date().isoformat()
                                except Exception:
                                    errores[fila]["bloqueantes"].append(f"fecha_invalida:{campo}")

                        # Si ambas fechas están presentes, comprobar coherencia: llegada >= extracción
                        if datos.get('fecha_extraccion') and datos.get('fecha_llegada'):
                            try:
                                fe_extr = date.fromisoformat(datos['fecha_extraccion'])
                                fe_lleg = date.fromisoformat(datos['fecha_llegada'])
                                if fe_lleg < fe_extr:
                                    errores[fila]["bloqueantes"].append('fecha_incoherente')
                            except Exception:
                                # Si por alguna razón no se pueden convertir, marcar como fecha inválida y bloquear
                                errores[fila]["bloqueantes"].append('fecha_invalida:fecha_extraccion')
                                    
                        # Comprobar si hay duplicados entre las muestras dentro del excel o en la base de datos
                        nom_lab = datos["nom_lab"]

                        if nom_lab in cache["muestras_existentes"]:
                            errores[fila]["bloqueantes"].append("muestra_duplicada_bd")
                        if nom_lab in nom_lab_excel:
                            errores[fila]["bloqueantes"].append("muestra_duplicada_excel")
                        else:
                            nom_lab_excel.add(nom_lab)
                        
                        # Comprobar si el estudio existe en la base de datos
                        estudio_id = datos.get("estudio")
                        if estudio_id:
                            try:
                                estudio_id = int(estudio_id)  # Convertir a integer para buscar en BD
                                estudio = cache["estudios"].get(estudio_id)
                                if not estudio:
                                    errores[fila]["advertencias"].append("estudio_no_existe")
                                    datos["estudio"] = None
                                else:
                                    datos["estudio"] = estudio.id
                            except (ValueError, TypeError):
                                errores[fila]["advertencias"].append("estudio_no_existe")
                                datos["estudio"] = None

                        # Comprobar si la localización está ocupada o existe únicamente si se han proporcionado datos de posición
                        if datos.get("congelador") or datos.get("estante") or datos.get("rack") or datos.get("caja") or datos.get("subposicion"):
                            key = (
                                datos["congelador"],
                                datos["estante"],
                                datos["rack"],
                                datos["caja"],
                                datos["subposicion"]
                            )

                            subposicion = cache["subposiciones"].get(key)

                            if not subposicion:
                                errores[fila]["bloqueantes"].append("localizacion_no_existe")
                            elif not subposicion.vacia:
                                errores[fila]["bloqueantes"].append("localizacion_ocupada")
                            else:
                                datos["subposicion_id"] = subposicion.id
                        else:
                            # No se proporcionó posición; la muestra quedará sin localización asignada
                            datos["subposicion_id"] = None

                        # Detectar campos opcionales vacios (solo 'nom_lab' es obligatorio)
                        opcionales = [
                            'id_individuo',
                            'id_material',
                            'volumen_actual',
                            'unidad_volumen',
                            'concentracion_actual',
                            'unidad_concentracion',
                            'masa_actual',
                            'unidad_masa',
                            'fecha_extraccion',
                            'fecha_llegada',
                            'observaciones',
                            'estado_inicial',
                            'estado_actual',
                            'estudio',
                            'centro_procedencia',
                            'lugar_procedencia',
                            # Posición (tratar como opcional; si se proporcionan, se validan arriba)
                            'congelador',
                            'estante',
                            'posicion_rack_estante',
                            'rack',
                            'posicion_caja_rack',
                            'caja',
                            'subposicion'
                        ]
                        for campo in opcionales:
                            if datos.get(campo) is None:
                                errores[fila]["advertencias"].append(f"campo_vacio:{campo}")

                        # Registrar filas validas
                        if not errores[fila]["bloqueantes"]:
                            datos['fila'] = fila
                            filas_validas.append(datos)

                        if _should_update(idx, total_filas):
                            yield _progress_update(idx + 1, total_filas)
                    
                    # Guardar en la sesión las filas validas y los errores detectados
                    request.session['filas_validas']=filas_validas
                    request.session['errores'] = errores

                    # Obtener configuración de mensajes para muestras
                    msg_config = get_upload_messages('muestras')
                    
                    # Contar errores
                    numero_errores_bloqueantes = 0
                    numero_errores_advertencia = 0
                    for fila in errores:
                        if errores[fila]['bloqueantes']:
                            numero_errores_bloqueantes += 1
                        if errores[fila]["advertencias"]:
                            numero_errores_advertencia += 1

                    # Construir lista de mensajes para confirmación
                    mensajes = []
                    mensajes.append({'level': 'info', 'text': f'{msg_config["titulo_inicial"]} {numero_registros} registros.'})

                    if numero_errores_bloqueantes == 0 and numero_errores_advertencia == 0:
                        mensajes.append({'level': 'success', 'text': msg_config['sin_errores']})
                    else:
                        if numero_errores_advertencia > 0:
                            msg = msg_config['con_advertencias'].format(count=numero_errores_advertencia)
                            mensajes.append({'level': 'warning', 'text': msg})
                        if numero_errores_bloqueantes > 0:
                            msg = msg_config['con_bloqueantes'].format(count=numero_errores_bloqueantes)
                            mensajes.append({'level': 'error', 'text': msg})
                    
                    # Mostrar mensaje de columnas extras si existen
                    columnas_extras_str = request.session.get('columnas_adicionales', '')
                    tiene_columnas_extras = bool(columnas_extras_str)
                    numero_columnas_extras = len(columnas_extras_str.split(', ')) if columnas_extras_str else 0
                    if tiene_columnas_extras:
                        msg = msg_config['columnas_extras'].format(count=numero_columnas_extras, detalles=columnas_extras_str)
                        mensajes.append({'level': 'warning', 'text': msg})

                    # Guardar datos de confirmación en sesión para mostrar tras redirección
                    request.session['confirmacion_pendiente'] = {
                        'template': 'confirmacion_upload.html',
                        'context': {
                            'numero_errores_bloqueantes': numero_errores_bloqueantes,
                            'numero_errores_advertencia': numero_errores_advertencia,
                            'tiene_columnas_extras': tiene_columnas_extras,
                            'numero_columnas_extras': numero_columnas_extras,
                            'columnas_extras_str': columnas_extras_str
                        },
                        'mensajes': mensajes
                    }
                    request.session.save()

                    yield _progress_done(request.path + '?mostrar_confirmacion=1', 'Validación completada')

                return StreamingHttpResponse(gen_validar_muestras(), content_type='text/html')
        # Si se solicita un excel de errores, este se rellena en base a los errores detectados durante la validación 
        elif 'excel_errores' in request.POST:
                    # Leer los errores y el excel de la sesión
                    errores = request.session.get('errores',[])
                    columnas_adicionales_str = request.session.get('columnas_adicionales', '')
                    # Convertir string de columnas adicionales a un set para procesamiento
                    columnas_adicionales = set()
                    if columnas_adicionales_str:
                        columnas_adicionales = set(col.strip() for col in columnas_adicionales_str.split(','))
                    
                    excel_bytes = base64.b64decode(request.session.get('excel_file_base64'))
                    excel_file = io.BytesIO(excel_bytes)
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    # Definir los estilos para pintar el excel usando configuración centralizada
                    colors = get_excel_colors()
                    FILL_ERROR_CELL = PatternFill("solid", fgColor=colors['error_cell'])
                    FILL_WARN_CELL  = PatternFill("solid", fgColor=colors['warning_cell'])
                    FILL_EXTRA_COL  = PatternFill("solid", fgColor=colors['extra_column'])
                    # Diccionario de mensajes
                    MENSAJES_ERROR = {
                        "campo_obligatorio_vacio": "Campo obligatorio vacío",
                        "formato_incorrecto": "Formato incorrecto",
                        "valor_negativo": "El valor debe ser positivo (≥0)",
                        "estado_actual_invalido": "Estado debe ser: Disponible, Enviada, Enviada parcialmente o Destruida",
                        "fecha_invalida": "Fecha inválida (Formato correcto: DD-MM-AAAA)",
                        "fecha_incoherente": "Fecha llegada anterior a fecha de extracción",
                        "muestra_duplicada_bd": "La muestra ya existe en la base de datos",
                        "muestra_duplicada_excel": "Muestra duplicada dentro del Excel",
                        "localizacion_ocupada": "La subposición ya está ocupada",
                        "localizacion_no_existe": "La localización no existe",
                        "campo_vacio": "Campo opcional vacío",
                        "estudio_no_existe": "El estudio no existe (se puede asignar después)",
                        "caracter_invalido_semicolon": "El carácter ';' no está permitido en este campo",
                    }
                    # Diccionario de columnas del excel
                    columnas_excel = {}
                    rename_columns = {
                        'ID Individuo': 'id_individuo', 
                        'Nombre Laboratorio': 'nom_lab',
                        'ID Material': 'id_material',
                        'Volumen Actual': 'volumen_actual',
                        'Unidad Volumen': 'unidad_volumen',
                        'Concentracion Actual': 'concentracion_actual',
                        'Unidad Concentracion': 'unidad_concentracion',
                        'Masa Actual': 'masa_actual',
                        'Unidad Masa': 'unidad_masa',
                        'Fecha Extraccion': 'fecha_extraccion',
                        'Fecha Llegada': 'fecha_llegada',
                        'Observaciones': 'observaciones',
                        'Estado Inicial': 'estado_inicial',
                        'Centro Procedencia': 'centro_procedencia',
                        'Lugar Procedencia': 'lugar_procedencia',
                        'Estado actual': 'estado_actual',
                        'Congelador': 'congelador', 
                        'Estante': 'estante',
                        'Posición del rack en el estante': 'posicion_rack_estante',
                        'Rack': 'rack',
                        'Posición de la caja en el rack': 'posicion_caja_rack',
                        'Caja': 'caja',
                        'Subposición': 'subposicion',
                        'Estudio':'estudio'
                    }
                    for cell in ws[1]:
                        if cell.value in rename_columns:
                            columnas_excel[rename_columns[cell.value]] = cell.column
                    # Añadir la columna de errores
                    col_errores = ws.max_column + 1
                    ws.cell(row=1, column=col_errores, value="Errores")
                    # Mapear errores sin campo a sus columnas específicas (para muestras)
                    error_campo_map = {
                        "muestra_duplicada_bd": "nom_lab",
                        "muestra_duplicada_excel": "nom_lab",
                        "localizacion_no_existe": "subposicion",
                        "localizacion_ocupada": "subposicion"
                    }
                    # Recorrer filas con errores 
                    for fila, info in errores.items():
                        has_error = bool(info.get("bloqueantes", []))
                        has_warn = bool(info.get("advertencias", []))
                        if not has_error and not has_warn:
                            continue

                        # Colorear celdas específicas y construir mensajes
                        mensajes = []
                        for err in info.get("bloqueantes", []):
                            if ":" in err:
                                tipo, campo = err.split(":")
                                msg = f"(Error) {MENSAJES_ERROR[tipo]}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                                if campo in columnas_excel:
                                    col = columnas_excel[campo]
                                    celda = ws.cell(row=int(fila), column=col)
                                    celda.fill = FILL_ERROR_CELL
                            else:
                                campo = error_campo_map.get(err)
                                msg = f"(Error) {MENSAJES_ERROR[err]}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                                if campo and campo in columnas_excel:
                                    col_err = columnas_excel[campo]
                                    ws.cell(row=int(fila), column=col_err).fill = FILL_ERROR_CELL
                        for warn in info.get("advertencias", []):
                            if ":" in warn:
                                tipo, campo = warn.split(":")
                                mensaje_warn = MENSAJES_ERROR.get(tipo, tipo)
                                msg = f"(Advertencia) {mensaje_warn}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                                if campo in columnas_excel:
                                    col = columnas_excel[campo]
                                    celda = ws.cell(row=int(fila), column=col)
                                    celda.fill = FILL_WARN_CELL
                            else:
                                mensaje_warn = MENSAJES_ERROR.get(warn, warn)
                                msg = f"(Advertencia) {mensaje_warn}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                        ws.cell(row=int(fila), column=col_errores, value="\n".join(mensajes))

                    # Pintar columnas extras con color de columna extra
                    if columnas_adicionales:
                        for col_name in columnas_adicionales:
                            # Encontrar el número de columna del Excel original para esta columna extra
                            for cell in ws[1]:
                                if cell.value == col_name:
                                    col_num = cell.column
                                    # Pintar el encabezado
                                    header_cell = ws.cell(row=1, column=col_num)
                                    header_cell.fill = FILL_EXTRA_COL
                                    # Pintar todas las celdas de datos en la columna
                                    for row in range(2, ws.max_row + 1):
                                        ws.cell(row=row, column=col_num).fill = FILL_EXTRA_COL
                                    break

                    # Rertornar el excel de errores    
                    output = io.BytesIO()    
                    wb.save(output)
                    wb.close()
                    response = HttpResponse(output.getvalue(),content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="listado_errores.xlsx"'
                    return response        
 
    else:
        form = UploadExcel(request)     
    return render(request, 'upload_excel.html', {'form': form}) 
@permission_required('muestras.can_change_muestras_web')
def cambio_posicion(request):
    # Mostrar confirmación después de validación con progreso
    if request.GET.get('mostrar_confirmacion') and 'confirmacion_pendiente' in request.session:
        datos_conf = request.session.pop('confirmacion_pendiente')
        for msg in datos_conf.get('mensajes', []):
            getattr(messages, msg['level'])(request, msg['text'])
        return render(request, datos_conf['template'], datos_conf.get('context', {}))
    # Vista para cambiar la posición de múltiples muestras a partir de un archivo Excel, requiere permiso para cambiar muestras
    if request.method=="POST":
        form = UploadExcel(request.POST, request.FILES)
        # Si el usuario confirma, se guardan las muestras en una nueva posición, vaciando la posicion antigua
        if 'confirmar' in request.POST:
            filas_validas = request.session.get('filas_validas',[])
            total = len(filas_validas)

            def gen_confirmar_cambio():
                yield _progress_page_start('Cambiando posiciones', total)
                try:
                    with transaction.atomic():
                        for i, datos in enumerate(filas_validas):
                            subposicion = Subposicion.objects.select_for_update().get(id=datos["subposicion_id"])
                            muestra = Muestra.objects.get(nom_lab=datos['nom_lab'])

                            localizacion = Localizacion.objects.create(
                                muestra=muestra,
                                congelador=subposicion.caja.rack.estante.congelador.congelador,
                                estante=subposicion.caja.rack.estante.numero,
                                rack=subposicion.caja.rack.numero,
                                caja=subposicion.caja.numero,
                                subposicion=subposicion.numero,
                            )

                            antigua_id = datos.get('subposicion_antigua')
                            if antigua_id:
                                try:
                                    subposicion_antigua = Subposicion.objects.select_for_update().get(id=antigua_id)
                                    subposicion_antigua.vacia = True
                                    subposicion_antigua.muestra = None
                                    subposicion_antigua.save()
                                except Subposicion.DoesNotExist:
                                    pass

                            subposicion.vacia = False
                            subposicion.muestra = muestra
                            subposicion.save()

                            historial_localizaciones.objects.create(
                                muestra=muestra,
                                localizacion=localizacion,
                                fecha_asignacion=timezone.now(),
                                usuario_asignacion=request.user
                            )
                            if _should_update(i, total):
                                yield _progress_update(i + 1, total)
                    yield _progress_done('/muestras/', 'Posiciones actualizadas correctamente')
                except Exception as e:
                    yield _progress_error(str(e))

            return StreamingHttpResponse(gen_confirmar_cambio(), content_type='text/html')

        # Si el usuario cancela, no se hace nada
        elif 'cancelar' in request.POST:
            messages.error(request,'Las muestras no se han cambiado de posición.')
            return redirect('muestras_todas')
        
        elif 'excel_file' in request.FILES:
            # Limpiar sesión residual de uploads anteriores
            if 'columnas_adicionales' in request.session:
                del request.session['columnas_adicionales']
            if form.is_valid():
                # Leer excel y preparar columnas 
                excel_file = request.FILES['excel_file']
                
                # Validar que sea un archivo Excel
                if not excel_file.name.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                    return render(request, 'upload_excel_cambio_posicion.html', {'form': form, 'error': '❌ Error de formato: El archivo debe ser un Excel (.xlsx, .xls o .xlsm).'})
                
                excel_bytes = excel_file.read()
                request.session['excel_file_name'] = excel_file.name
                request.session['excel_file_base64']= base64.b64encode(excel_bytes).decode()
                excel_stream = io.BytesIO(excel_bytes)
                
                # Intentar leer el archivo Excel
                try:
                    df = pd.read_excel(excel_stream)
                except Exception as e:
                    return render(request, 'upload_excel_cambio_posicion.html', {'form': form, 'error': f'❌ Error al leer el archivo Excel: {str(e)}'})
                
                rename_columns = {
                    'Nombre Laboratorio': 'nom_lab',
                    'Congelador': 'congelador', 
                    'Estante': 'estante',
                    'Posición del rack en el estante': 'posicion_rack_estante',
                    'Rack': 'rack',
                    'Posición de la caja en el rack': 'posicion_caja_rack',
                    'Caja': 'caja',
                    'Subposición': 'subposicion',
                }
                
                # Normalizar nombres de columnas a minúsculas para comparación insensible a mayúsculas/minúsculas
                df.columns = df.columns.str.lower()
                rename_columns_normalized = {k.lower(): v for k, v in rename_columns.items()}
                
                # Validar que el Excel tenga las columnas esperadas
                columnas_esperadas = set(rename_columns_normalized.keys())
                columnas_existentes = set(df.columns)
                columnas_faltantes = columnas_esperadas - columnas_existentes
                
                if columnas_faltantes:
                    columnas_str = ', '.join(sorted(columnas_faltantes))
                    return render(request, 'upload_excel_cambio_posicion.html', {'form': form, 'error': f'❌ Error de formato: El archivo Excel no contiene las siguientes columnas esperadas: {columnas_str}'})
                
                # Validar que el Excel no esté vacío
                if df.empty or len(df) == 0:
                    return render(request, 'upload_excel_cambio_posicion.html', {'form': form, 'error': '❌ Error de formato: El archivo Excel está vacío o no contiene filas de datos.'})
                
                # Validar columnas adicionales
                columnas_adicionales = columnas_existentes - columnas_esperadas
                extra_columns = False
                columnas_adicionales_str = ''
                if columnas_adicionales:
                    columnas_adicionales_str = ', '.join(sorted(columnas_adicionales))
                    request.session['columnas_adicionales'] = columnas_adicionales_str
                    extra_columns = True
                
                df.rename(columns=rename_columns_normalized, inplace=True)
                # Funciones para normalizar las columnas del excel
                def norm(value):
                    if value is None or pd.isna(value):
                        return None

                    if isinstance(value, str):
                        value = value.strip()
                        return value if value != "" else None

                    return value
                
                def norm_code(value):
                    if value is None or pd.isna(value):
                        return None

                    if isinstance(value, float) and value.is_integer():
                        return str(int(value)).lower()

                    return str(value).strip().lower()
                # Carga de datos previos y creación de estructuras 
                cache = {
                    'subposiciones': {
                        (
                            str(c.caja.rack.estante.congelador.congelador).lower(),
                            c.caja.rack.estante.numero,
                            str(c.caja.rack.numero).lower(),
                            str(c.caja.numero).lower(),
                            str(c.numero).lower()
                        ): c
                         for c in Subposicion.objects.select_related('caja__rack__estante__congelador')
                    },
                    'posiciones_actuales': {
                        (p.muestra.nom_lab.lower()): p.id 
                        for p in Subposicion.objects.all() if p.muestra != None
                    },

                    'muestras_existentes': set(m.lower() for m in Muestra.objects.values_list('nom_lab',flat=True))
                }

                total_filas = len(df)

                def gen_validar_cambio():
                    filas_validas = []
                    errores = {}
                    nom_lab_excel = set()
                    numero_registros = 0

                    yield _progress_page_start('Validando cambio de posición', total_filas)

                    # Recorrer el df para detectar errores y normalizar
                    for idx, row in df.iterrows():
                        numero_registros += 1
                        fila = idx + 2 
                        errores[fila]={"bloqueantes":[]}
                        datos = {
                            "nom_lab":norm(row['nom_lab']),
                            "congelador":(norm_code(row['congelador']) or None),
                            "estante":norm_code(row['estante']),
                            "posicion_rack_estante":norm_code(row['posicion_rack_estante']),
                            "rack":(norm_code(row['rack']) or None),
                            "posicion_caja_rack":norm_code(row['posicion_caja_rack']),
                            "caja":(norm_code(row['caja']) or None),
                            "subposicion":(norm_code(row['subposicion']) or None), 
                        }
                        # Comprobar si los campos obligatorios están rellenados
                        obligatorios = ["nom_lab", "congelador", "estante", "posicion_rack_estante", "rack", "caja", "posicion_caja_rack","subposicion"]

                        for campo in obligatorios:
                            if not datos.get(campo):
                                errores[fila]["bloqueantes"].append(f"campo_obligatorio_vacio:{campo}")
                        
                        # Comprobar si hay duplicados entre las muestras dentro del excel o si la muestra no está en la base de datos
                        nom_lab = datos["nom_lab"].lower()

                        if nom_lab not in cache["muestras_existentes"]:
                            errores[fila]["bloqueantes"].append("muestra_no_existe_bd")
                        if nom_lab in nom_lab_excel:
                            errores[fila]["bloqueantes"].append("muestra_duplicada_excel")
                        else:
                            nom_lab_excel.add(nom_lab)

                        # Comprobar si la localización está ocupada o existe
                        key = (
                            datos["congelador"],
                            datos["estante"],
                            datos["rack"],
                            datos["caja"],
                            datos["subposicion"]
                        )

                        subposicion = cache["subposiciones"].get(key)

                        if not subposicion:
                            errores[fila]["bloqueantes"].append("localizacion_no_existe")
                        elif not subposicion.vacia:
                            errores[fila]["bloqueantes"].append("localizacion_ocupada")
                        else:
                            datos["subposicion_id"] = subposicion.id
                            datos["subposicion_antigua"] = cache['posiciones_actuales'].get(datos['nom_lab'].lower())
                        
                        # Registrar filas validas
                        if not errores[fila]["bloqueantes"]:
                            datos['fila'] = fila
                            filas_validas.append(datos)

                        if _should_update(idx, total_filas):
                            yield _progress_update(idx + 1, total_filas)

                    # Guardar en la sesión las filas validas y los errores detectados
                    request.session['filas_validas'] = filas_validas
                    request.session['errores'] = errores

                    # Mensajes de información de la subida
                    upload_msgs = get_upload_messages('cambio_posicion')
                    mensajes = [{'level': 'info', 'text': f"{upload_msgs['titulo_inicial']} {numero_registros} registros."}]
                    numero_errores_bloqueantes = 0
                    for fila in errores:
                        if errores[fila]['bloqueantes']:
                            numero_errores_bloqueantes += 1
                    
                    errores_encontrados = numero_errores_bloqueantes > 0 or extra_columns
                    
                    if numero_errores_bloqueantes == 0 and not extra_columns:
                        mensajes.append({'level': 'success', 'text': upload_msgs['sin_errores']})
                    else:
                        if numero_errores_bloqueantes > 0:
                            msg_bloqueantes = upload_msgs['con_bloqueantes'].format(count=numero_errores_bloqueantes)
                            mensajes.append({'level': 'error', 'text': msg_bloqueantes})
                        if extra_columns:
                            msg_extras = upload_msgs['columnas_extras'].format(count=len(columnas_adicionales), detalles=columnas_adicionales_str)
                            mensajes.append({'level': 'warning', 'text': msg_extras})

                    # Guardar datos de confirmación en sesión para mostrar tras redirección
                    request.session['confirmacion_pendiente'] = {
                        'template': 'confirmacion_upload_cambio_posicion.html',
                        'context': {'errores_encontrados': errores_encontrados},
                        'mensajes': mensajes
                    }
                    request.session.save()
                    yield _progress_done(request.path + '?mostrar_confirmacion=1', 'Validación completada')

                return StreamingHttpResponse(gen_validar_cambio(), content_type='text/html')
            
        # Si se solicita un excel de errores, este se rellena en base a los errores detectados durante la validación 
        elif 'excel_errores' in request.POST:
                    # Leer los errores y el excel de la sesión
                    errores = request.session.get('errores',[])
                    columnas_adicionales_str = request.session.get('columnas_adicionales', '')
                    columnas_adicionales = set()
                    if columnas_adicionales_str:
                        columnas_adicionales = set(col.strip() for col in columnas_adicionales_str.split(','))
                    
                    excel_bytes = base64.b64decode(request.session.get('excel_file_base64'))
                    excel_file = io.BytesIO(excel_bytes)
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    # Definir los estilos para pintar el excel usando configuración centralizada
                    colors = get_excel_colors()
                    FILL_ERROR_CELL = PatternFill("solid", fgColor=colors['error_cell'])
                    FILL_EXTRA_COL  = PatternFill("solid", fgColor=colors['extra_column'])
                    # Diccionario de mensajes
                    MENSAJES_ERROR = {
                        "campo_obligatorio_vacio": "Campo obligatorio vacío",
                        "muestra_no_existe_bd": "La muestra no existe en la base de datos",
                        "muestra_duplicada_excel": "Muestra duplicada dentro del Excel",
                        "localizacion_ocupada": "La subposición ya está ocupada",
                        "localizacion_no_existe": "La localización no existe",
                        "fecha_invalida": "Fecha inválida (Formato correcto: DD-MM-AAAA)",
                        "caracter_invalido_semicolon": "El carácter ';' no está permitido en este campo"
                    }
                    # Diccionario de columnas del excel
                    columnas_excel = {}
                    rename_columns = {
                        'Nombre Laboratorio': 'nom_lab',
                        'Congelador': 'congelador', 
                        'Estante': 'estante',
                        'Posición del rack en el estante': 'posicion_rack_estante',
                        'Rack': 'rack',
                        'Posición de la caja en el rack': 'posicion_caja_rack',
                        'Caja': 'caja',
                        'Subposición': 'subposicion',
                    }
                    for cell in ws[1]:
                        if cell.value in rename_columns:
                            columnas_excel[rename_columns[cell.value]] = cell.column
                    # Añadir la columna de errores
                    col_errores = ws.max_column + 1
                    ws.cell(row=1, column=col_errores, value="Errores")
                    # Mapear errores sin campo a sus columnas específicas (cambio posición)
                    error_campo_map_cambio = {
                        "muestra_no_existe_bd": "nom_lab",
                        "muestra_duplicada_excel": "nom_lab",
                        "localizacion_ocupada": "subposicion",
                        "localizacion_no_existe": "congelador"
                    }
                    # Recorrer filas con errores 
                    for fila, info in errores.items():
                        has_error = bool(info["bloqueantes"])
                        if not has_error:
                            continue

                        # Colorear celdas específicas y construir mensajes
                        mensajes = []
                        for err in info["bloqueantes"]:
                            if ":" in err:
                                tipo, campo = err.split(":")
                                msg = f"(Error) {MENSAJES_ERROR[tipo]}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                                col = columnas_excel[campo]
                                celda = ws.cell(row=int(fila), column=col)
                                celda.fill = FILL_ERROR_CELL
                            else:
                                campo = error_campo_map_cambio.get(err)
                                msg = f"(Error) {MENSAJES_ERROR[err]}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                                if campo and campo in columnas_excel:
                                    col_err = columnas_excel[campo]
                                    ws.cell(row=int(fila), column=col_err).fill = FILL_ERROR_CELL
                        ws.cell(row=int(fila), column=col_errores, value="\n".join(mensajes))

                    # Pintar columnas extras con color de columna extra
                    if columnas_adicionales:
                        # Comparación case-insensitive porque la validación normaliza a minúsculas
                        extras_lower = {c.lower() for c in columnas_adicionales}
                        for cell in ws[1]:
                            if cell.value and str(cell.value).lower() in extras_lower:
                                col_num = cell.column
                                # Pintar el encabezado
                                ws.cell(row=1, column=col_num).fill = FILL_EXTRA_COL
                                # Pintar todas las celdas de datos en la columna
                                for row in range(2, ws.max_row + 1):
                                    ws.cell(row=row, column=col_num).fill = FILL_EXTRA_COL

                    # Rertornar el excel de errores    
                    output = io.BytesIO()    
                    wb.save(output)
                    wb.close()
                    response = HttpResponse(output.getvalue(),content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="listado_errores.xlsx"'
                    return response        
    else:
        form = UploadExcel(request)
    return render(request, 'upload_excel_cambio_posicion.html', {'form': form}) 

@permission_required('muestras.can_change_muestras_web')
def editar_muestra(request, nom_lab):
    # Vista para editar una muestra existente, requiere permiso para cambiar muestras
    muestra = Muestra.objects.get(nom_lab=nom_lab)
    if request.method == 'POST':
        # Guardar el estudio anterior ANTES de is_valid(), ya que is_valid() modifica la instancia
        estudio_anterior = muestra.estudio
        form = MuestraForm(request.POST, instance=muestra)
        if form.is_valid():
            # Verificar si nom_lab ha sido cambiado
            nom_lab_anterior = muestra.nom_lab
            nom_lab_nuevo = form.cleaned_data.get('nom_lab')
            
            with connection.cursor() as cursor:
                # Desactivar las restricciones de clave foránea
                cursor.execute("SET FOREIGN_KEY_CHECKS=0")
            
            try:
                with transaction.atomic():
                    # Si nom_lab cambió y hay localizaciones, actualizar todas las referencias
                    if nom_lab_nuevo != nom_lab_anterior and muestra.localizacion.exists():
                        with connection.cursor() as cursor:
                            # Actualizar todas las localizaciones con el nuevo nom_lab
                            cursor.execute(
                                "UPDATE muestras_localizacion SET muestra_id = %s WHERE muestra_id = %s",
                                [nom_lab_nuevo, nom_lab_anterior]
                            )
                    
                    # Preservar fechas si el usuario las dejó vacías pero había valores anteriores
                    muestra_guardada = form.save(commit=False)
                    if not form.cleaned_data.get('fecha_extraccion') and muestra.fecha_extraccion:
                        muestra_guardada.fecha_extraccion = muestra.fecha_extraccion
                    if not form.cleaned_data.get('fecha_llegada') and muestra.fecha_llegada:
                        muestra_guardada.fecha_llegada = muestra.fecha_llegada
                    muestra_guardada.save()
                    
                    # Registrar en historial de estudios si el estudio ha cambiado
                    estudio_nuevo = muestra_guardada.estudio
                    if estudio_nuevo != estudio_anterior:
                        historial_estudios.objects.create(
                            muestra=muestra_guardada,
                            estudio=estudio_nuevo,
                            fecha_asignacion=timezone.now(),
                            usuario_asignacion=request.user
                        )
                    
                    # Manejar cambio de localización si se proporcionó
                    subposicion_id = request.POST.get("subposicion_id")
                    if subposicion_id:
                        try:
                            subposicion = Subposicion.objects.select_for_update().get(id=subposicion_id)
                            
                            if subposicion.vacia:
                                # Vaciar la subposición antigua si existe
                                if Subposicion.objects.filter(muestra=muestra).exists():
                                    subposicion_antigua = Subposicion.objects.select_for_update().get(muestra=muestra)
                                    subposicion_antigua.muestra = None
                                    subposicion_antigua.vacia = True
                                    subposicion_antigua.save()

                                # Asignar la nueva subposición
                                subposicion.muestra = muestra
                                subposicion.vacia = False
                                subposicion.save()

                                # Crear localización
                                localizacion = Localizacion.objects.create(
                                    muestra=muestra,
                                    congelador=subposicion.caja.rack.estante.congelador.congelador,
                                    estante=subposicion.caja.rack.estante.numero,
                                    rack=subposicion.caja.rack.numero,
                                    caja=subposicion.caja.numero,
                                    subposicion=subposicion.numero,
                                )

                                historial_localizaciones.objects.create(
                                    muestra=muestra,
                                    localizacion=localizacion,
                                    fecha_asignacion=timezone.now(),
                                    usuario_asignacion=request.user
                                )
                                messages.success(request, 'Muestra editada correctamente')
                            else:
                                messages.error(request, 'La subposición está ocupada por otra muestra.')
                        except Subposicion.DoesNotExist:
                            messages.error(request, 'La subposición seleccionada no existe.')
                    else:
                        messages.success(request, 'Muestra editada correctamente')
            finally:
                # Reactivar las restricciones de clave foránea
                with connection.cursor() as cursor:
                    cursor.execute("SET FOREIGN_KEY_CHECKS=1")
            
            return redirect('muestras_todas')
    else:
        form = MuestraForm(instance=muestra)
    
    # Obtener datos para los menús desplegables
    congeladores = Congelador.objects.all().values('id', 'congelador').order_by('congelador')
    
    # Obtener localización actual si existe. No acceder directamente a
    # `muestra.subposicion` porque para relaciones OneToOneDescriptor el acceso
    # puede lanzar RelatedObjectDoesNotExist cuando no existe la relación.
    localizacion_actual = None
    subposicion_actual = Subposicion.objects.filter(muestra=muestra).select_related(
        'caja__rack__estante__congelador'
    ).first()
    if subposicion_actual:
        localizacion_actual = {
            'congelador_id': subposicion_actual.caja.rack.estante.congelador.id,
            'congelador_nombre': subposicion_actual.caja.rack.estante.congelador.congelador,
            'estante_id': subposicion_actual.caja.rack.estante.id,
            'estante_numero': subposicion_actual.caja.rack.estante.numero,
            'rack_id': subposicion_actual.caja.rack.id,
            'rack_numero': subposicion_actual.caja.rack.numero,
            'caja_id': subposicion_actual.caja.id,
            'caja_numero': subposicion_actual.caja.numero,
            'subposicion_id': subposicion_actual.id,
            'subposicion_numero': subposicion_actual.numero,
        }
    
    context = {
        'form': form,
        'muestra': muestra,
        'congeladores': congeladores,
        'localizacion_actual': localizacion_actual,
        'localizacion_actual_json': json.dumps(localizacion_actual),
    }
    
    return render(request, 'editar_muestra.html', context)

def descargar_plantilla(request,macro:int):
    # Vista para descargar la plantilla de Excel para subir localizaciones o muestras
    if macro == 0:
        plantilla_path = os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_localizaciones.xlsx')
        if os.path.exists(plantilla_path):
            return FileResponse(open(plantilla_path, 'rb'), as_attachment=True, filename='plantilla_localizaciones.xlsx')
    elif macro == 1:
        plantilla_path = os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_localizaciones_macros.xlsm')
        if os.path.exists(plantilla_path):
            return FileResponse(open(plantilla_path, 'rb'), as_attachment=True, filename='plantilla_localizaciones_macros.xlsm')
    elif macro == 2:
        plantilla_path = os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_muestras.xlsx')
        if os.path.exists(plantilla_path):
            return FileResponse(open(plantilla_path, 'rb'), as_attachment=True, filename='plantilla_muestras.xlsx')
    elif macro == 3:
        plantilla_path = os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_estudios.xlsx')
        if os.path.exists(plantilla_path):
            return FileResponse(open(plantilla_path, 'rb'), as_attachment=True, filename='plantilla_estudios.xlsx')
    elif macro == 4:
        plantilla_path = os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_cambio_posicion.xlsx')
        if os.path.exists(plantilla_path):
            return FileResponse(open(plantilla_path, 'rb'), as_attachment=True, filename='plantilla_cambio_posicion.xlsx')
    else:
        return HttpResponse("La plantilla no se encuentra disponible.", status=404)
    
# Vistas para Localizacion
@login_required
@permission_required('muestras.can_view_localizaciones_web')
def localizaciones(request):
    # Vista que muestra todas las localizaciones, tengan o no muestra

    # Anotar el número de muestras en cada caja
    cajas_qs = Caja.objects.annotate(
        numero_muestras=Count(
            'subposiciones',
            filter=Q(subposiciones__vacia=False)
        )
    )

    # Prefetch para optimizar consultas (sin subposiciones, se cargan por AJAX)
    congeladores = Congelador.objects.prefetch_related(
        Prefetch(
            'estantes__racks__cajas',
            queryset=cajas_qs
        )
    )


    

    template = loader.get_template('localizaciones_todas.html')

    context = {
        'congeladores':congeladores,
    }
    return HttpResponse(template.render(context, request))
@permission_required('muestras.can_add_localizaciones_web')
def upload_excel_localizaciones(request):
    # Mostrar confirmación después de validación con progreso
    if request.GET.get('mostrar_confirmacion') and 'confirmacion_pendiente' in request.session:
        datos_conf = request.session.pop('confirmacion_pendiente')
        for msg in datos_conf.get('mensajes', []):
            getattr(messages, msg['level'])(request, msg['text'])
        return render(request, datos_conf['template'], datos_conf.get('context', {}))
    if request.method=="POST":
        # Vista para subir localizaciones desde un archivo Excel, requiere permiso para añadir localizaciones
        form = UploadExcel(request.POST, request.FILES)
        # Definir el mapeo de columnas
        rename_columns = {
                    'Congelador': 'congelador', 
                    'Estante': 'estante',
                    'Posición del rack en el estante': 'posicion_rack_estante',
                    'Rack': 'rack',
                    'Posición de la caja en el rack': 'posicion_caja_rack',
                    'Caja': 'caja',
                    'Subposición': 'subposicion'
                }
        # Función para limpiar y normalizar los valores
        def limpiar_numero(valor):
            if pd.isna(valor):
                return None

            # Normalizar floats tipo 1.0 → 1
            if isinstance(valor, float) and valor.is_integer():
                valor = int(valor)

            valor = str(valor).strip()

            if valor == "":
                return None

            return valor

        # Normalizar campos de texto para tratar mayúsculas/minúsculas de la misma forma
        def normalizar_texto(valor):
            v = limpiar_numero(valor)
            if v is None:
                return None
            try:
                return str(v).upper()
            except Exception:
                return str(v)

        # Si el usuario confirma, se guardan las localizaciones en la base de datos      
        if 'confirmar' in request.POST:
            filas = request.session.pop('filas_validas', [])
            total = len(filas)

            def gen_confirmar_localizaciones():
                yield _progress_page_start('Importando localizaciones', total)
                try:
                    with transaction.atomic():
                        for i, fila in enumerate(filas):
                            congelador, _ = Congelador.objects.get_or_create(
                                congelador=fila['congelador']
                            )
                            estante, _ = Estante.objects.get_or_create(
                                congelador=congelador,
                                numero=fila['estante']
                            )
                            rack, _ = Rack.objects.get_or_create(
                                estante=estante,
                                numero=fila['rack'],
                                defaults={'posicion_rack_estante': fila['posicion_rack_estante']}
                            )
                            caja, _ = Caja.objects.get_or_create(
                                rack=rack,
                                numero=fila['caja'],
                                defaults={'posicion_caja_rack': fila['posicion_caja_rack']}
                            )
                            Subposicion.objects.get_or_create(
                                caja=caja,
                                numero=fila['subposicion']
                            )
                            if _should_update(i, total):
                                yield _progress_update(i + 1, total)
                    yield _progress_done('/archivo/', 'Localizaciones importadas correctamente')
                except Exception as e:
                    yield _progress_error(str(e))

            return StreamingHttpResponse(gen_confirmar_localizaciones(), content_type='text/html')
        # Si el usuario cancela, no se hace nada
        elif 'cancelar' in request.POST:
            messages.error(request, 'Las localizaciones del excel no se han añadido.')
            return redirect('localizaciones_todas')
        # Si se sube un archivo excel, se procesa y valida
        elif 'excel_file' in request.FILES:
            # Limpiar sesión residual de uploads anteriores
            if 'columnas_adicionales' in request.session:
                del request.session['columnas_adicionales']
            if form.is_valid():
                # Leer excel
                excel_file = request.FILES['excel_file']
                excel_bytes = excel_file.read()
                request.session['excel_file_base64']= base64.b64encode(excel_bytes).decode()
                excel_stream = io.BytesIO(excel_bytes)
                
                # Validar que sea un archivo Excel válido
                try:
                    df = pd.read_excel(excel_stream)
                except Exception as e:
                    return render(request, 'localizacion_nueva.html', {'form': form, 'error': f'El archivo no es un Excel válido (.xlsx). Error: {str(e)}'})
                
                # Validar que tenga al menos una fila de datos
                if len(df) == 0:
                    return render(request, 'localizacion_nueva.html', {'form': form, 'error': 'Error de formato: El archivo no presenta registros'})
                
                # Validar que tenga todas las columnas esperadas
                columnas_esperadas = set(rename_columns.keys())
                columnas_existentes = set(df.columns)
                columnas_faltantes = columnas_esperadas - columnas_existentes
                
                if columnas_faltantes:
                    columnas_str = ', '.join(sorted(columnas_faltantes))
                    return render(request, 'localizacion_nueva.html', {'form': form, 'error': f'Error de formato: El archivo Excel no contiene las siguientes columnas esperadas: {columnas_str}'})
                
                # Validar que no haya columnas adicionales
                columnas_adicionales = columnas_existentes - columnas_esperadas
                extra_columns = False
                columnas_adicionales_str = ''
                if columnas_adicionales:
                    columnas_adicionales_str = ', '.join(sorted(columnas_adicionales))
                    # Registrar en la sesión el excel original para descarga y avisar mediante mensajes
                    request.session['excel_file_base64'] = base64.b64encode(excel_bytes).decode()
                    request.session['excel_file_name'] = excel_file.name
                    # Guardar columnas adicionales en sesión para que aparezcan en el Excel de errores
                    request.session['columnas_adicionales'] = columnas_adicionales_str
                    extra_columns = True

                # renombrar columnas
                df.rename(columns=rename_columns, inplace=True)
                
                # Procesar y validar filas
                total_filas = len(df)

                def gen_validar_localizaciones():
                    errores = {}
                    filas_validas = []
                    numero_registros = len(df)
                    # Mapas para comprobar consistencias:
                    # - que una misma posición de caja no tenga cajas diferentes
                    # - que una misma posición de rack no tenga racks diferentes
                    pos_to_caja = {}
                    pos_rack_to_rack = {}
                    # detectar si la misma subposición se usa más de una vez dentro del Excel
                    subposiciones_usadas = set()

                    yield _progress_page_start('Validando localizaciones', total_filas)

                    for idx, row in df.iterrows():
                        fila_numero = idx + 2
                        errores[fila_numero] = {"bloqueantes": []}
                        
                        # Limpiar y normalizar los valores
                        congelador = normalizar_texto(row['congelador'])
                        estante = limpiar_numero(row['estante'])
                        posicion_rack_estante = normalizar_texto(row['posicion_rack_estante'])
                        rack = normalizar_texto(row['rack'])
                        posicion_caja_rack = normalizar_texto(row['posicion_caja_rack'])
                        caja = normalizar_texto(row['caja'])
                        subpos = normalizar_texto(row['subposicion'])
                        
                        # Comprobar si hay campos vacíos
                        campos = {
                            'congelador': congelador,
                            'estante': estante,
                            'posicion_rack_estante': posicion_rack_estante,
                            'rack': rack,
                            'posicion_caja_rack': posicion_caja_rack,
                            'caja': caja,
                            'subposicion': subpos
                        }
                        
                        for nombre_campo, valor in campos.items():
                            if valor is None:
                                errores[fila_numero]["bloqueantes"].append(f"campo_obligatorio_vacio:{nombre_campo}")
                            # Validar que no haya punto y coma en ningún campo
                            for nombre_campo, valor in campos.items():
                                if valor and isinstance(valor, str) and ';' in valor:
                                    errores[fila_numero]["bloqueantes"].append(f"caracter_invalido_semicolon:{nombre_campo}")
                        
                        # Validar que ciertos campos numéricos sean enteros positivos (>0)
                        if not errores[fila_numero]["bloqueantes"]:
                            try:
                                if estante is not None:
                                    if int(estante) <= 0:
                                        raise ValueError()
                            except Exception:
                                errores[fila_numero]["bloqueantes"].append("formato_incorrecto:estante")

                            try:
                                if posicion_rack_estante is not None:
                                    if int(posicion_rack_estante) <= 0:
                                        raise ValueError()
                            except Exception:
                                errores[fila_numero]["bloqueantes"].append("formato_incorrecto:posicion_rack_estante")

                            try:
                                if posicion_caja_rack is not None:
                                    if int(posicion_caja_rack) <= 0:
                                        raise ValueError()
                            except Exception:
                                errores[fila_numero]["bloqueantes"].append("formato_incorrecto:posicion_caja_rack")

                        # Si hay errores bloqueantes hasta ahora, saltar validaciones posteriores
                        if errores[fila_numero]["bloqueantes"]:
                            if _should_update(idx, total_filas):
                                yield _progress_update(idx + 1, total_filas)
                            continue

                        # Comprobar consistencia de rack: misma (congelador, estante, posicion_rack_estante)
                        # no puede mapear a racks distintos
                        pos_rack_key = (congelador, estante, posicion_rack_estante)
                        if pos_rack_key in pos_rack_to_rack:
                            if pos_rack_to_rack[pos_rack_key] != rack:
                                errores[fila_numero]["bloqueantes"].append("rack_inconsistente")
                                if _should_update(idx, total_filas):
                                    yield _progress_update(idx + 1, total_filas)
                                continue
                        else:
                            pos_rack_to_rack[pos_rack_key] = rack

                        # Comprobar consistencia: una misma posición de caja no puede tener cajas distintas
                        pos_key = (congelador, estante, posicion_rack_estante, rack, posicion_caja_rack)
                        if pos_key in pos_to_caja:
                            if pos_to_caja[pos_key] != caja:
                                errores[fila_numero]["bloqueantes"].append("caja_inconsistente")
                                if _should_update(idx, total_filas):
                                    yield _progress_update(idx + 1, total_filas)
                                continue
                        else:
                            pos_to_caja[pos_key] = caja

                        # Comprobar duplicado dentro del Excel: misma subposición completa ya usada
                        subpos_key = (congelador, estante, posicion_rack_estante, rack, posicion_caja_rack, subpos)
                        if subpos_key in subposiciones_usadas:
                            errores[fila_numero]["bloqueantes"].append("subposicion_duplicada_excel")
                            if _should_update(idx, total_filas):
                                yield _progress_update(idx + 1, total_filas)
                            continue
                        else:
                            subposiciones_usadas.add(subpos_key)
                        
                        # Comprobar si la posición del rack ya está ocupada por otro rack
                        if Rack.objects.filter(
                            estante__congelador__congelador__iexact=congelador,
                            estante__numero__iexact=estante,
                            posicion_rack_estante__iexact=posicion_rack_estante
                        ).exclude(numero__iexact=rack).exists():
                            errores[fila_numero]["bloqueantes"].append("posicion_rack_ocupada")
                            if _should_update(idx, total_filas):
                                yield _progress_update(idx + 1, total_filas)
                            continue
                        
                        # Comprobar si la posición de la caja ya está ocupada por otra caja
                        if Caja.objects.filter(
                            rack__estante__congelador__congelador__iexact=congelador,
                            rack__estante__numero__iexact=estante,
                            rack__numero__iexact=rack,
                            posicion_caja_rack__iexact=posicion_caja_rack
                        ).exclude(numero__iexact=caja).exists():
                            errores[fila_numero]["bloqueantes"].append("posicion_caja_ocupada")
                            if _should_update(idx, total_filas):
                                yield _progress_update(idx + 1, total_filas)
                            continue
                        
                        # Comprobar si la localización ya existe (case-insensitive para textos)
                        if Subposicion.objects.filter(numero__iexact=subpos,
                                                      caja__numero__iexact=caja,
                                                      caja__rack__numero__iexact=rack,
                                                      caja__rack__estante__numero=estante,
                                                      caja__rack__estante__congelador__congelador__iexact=congelador).exists():
                            errores[fila_numero]["bloqueantes"].append("localizacion_duplicada")
                        else:
                            # Guardar fila válida
                            filas_validas.append({
                                'congelador': congelador,
                                'estante': estante,
                                'posicion_rack_estante': posicion_rack_estante,
                                'rack': rack,
                                'posicion_caja_rack': posicion_caja_rack,
                                'caja': caja,
                                'subposicion': subpos
                            })

                        if _should_update(idx, total_filas):
                            yield _progress_update(idx + 1, total_filas)

                    # Guardar en sesión los resultados de la validación
                    request.session['filas_validas'] = filas_validas
                    request.session['errores'] = errores

                    # Obtener configuración de mensajes para localizaciones
                    msg_config = get_upload_messages('localizaciones')

                    # Build messages
                    mensajes = [{'level': 'info', 'text': f'{msg_config["titulo_inicial"]} {numero_registros} registros'}]

                    # Contar errores
                    numero_errores_bloqueantes = sum(1 for fila in errores if errores[fila]['bloqueantes'])

                    # Determinar si hay errores para mostrar la sección de Excel de errores
                    errores_encontrados = (numero_errores_bloqueantes > 0) or extra_columns

                    # Generar mensajes según el estado
                    if numero_errores_bloqueantes > 0:
                        msg = msg_config['con_bloqueantes'].format(count=numero_errores_bloqueantes)
                        mensajes.append({'level': 'error', 'text': msg})

                    if extra_columns:
                        num_extras = len([c.strip() for c in columnas_adicionales_str.split(',') if c.strip()])
                        msg = msg_config['columnas_extras'].format(count=num_extras, detalles=columnas_adicionales_str)
                        mensajes.append({'level': 'warning', 'text': msg})

                    if numero_errores_bloqueantes == 0 and not extra_columns:
                        mensajes.append({'level': 'success', 'text': msg_config['sin_errores']})

                    request.session['confirmacion_pendiente'] = {
                        'template': 'confirmacion_upload_localizacion.html',
                        'context': {'errores_encontrados': errores_encontrados},
                        'mensajes': mensajes
                    }
                    request.session.save()
                    yield _progress_done(request.path + '?mostrar_confirmacion=1', 'Validación completada')

                return StreamingHttpResponse(gen_validar_localizaciones(), content_type='text/html')
            
        # Si se solicita un excel de errores, este se rellena en base a los errores detectados durante la validación
        elif 'excel_errores' in request.POST:
            errores = request.session.get('errores', {})
            excel_bytes = base64.b64decode(request.session.get('excel_file_base64'))
            excel_file = io.BytesIO(excel_bytes)
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Definir los estilos para pintar el excel usando configuración centralizada
            colors = get_excel_colors()
            FILL_ERROR_CELL = PatternFill("solid", fgColor=colors['error_cell'])
            FILL_ERROR_COL = PatternFill("solid", fgColor=colors['extra_column'])
            
            # Diccionario de mensajes
            MENSAJES_ERROR = {
                "campo_obligatorio_vacio": "Campo obligatorio vacío",
                "localizacion_duplicada": "La localización ya existe en la base de datos",
                "subposicion_duplicada_excel": "La subposición aparece duplicada en el Excel",
                "formato_incorrecto": "Formato incorrecto (debe ser entero positivo)",
                "fecha_invalida": "Fecha inválida (Formato correcto: DD-MM-AAAA)",
                "caja_inconsistente": "Conflicto de caja en la misma posición",
                "rack_inconsistente": "Conflicto de rack en la misma posición",
                "posicion_rack_ocupada": "La posición del rack ya está ocupada por otro rack",
                "posicion_caja_ocupada": "La posición de la caja ya está ocupada por otra caja",
                "caracter_invalido_semicolon": "El carácter ';' no está permitido en este campo"
            }
            
            # Diccionario de columnas del excel
            columnas_excel = {}
            for cell in ws[1]:
                if cell.value in rename_columns:
                    columnas_excel[rename_columns[cell.value]] = cell.column
            
            # Añadir la columna de errores
            col_errores = ws.max_column + 1
            ws.cell(row=1, column=col_errores, value="Errores")
            
            # Comprobar si hay columnas adicionales registradas en sesión
            columnas_adicionales_str = request.session.get('columnas_adicionales', '')
            extra_columns_flag = bool(columnas_adicionales_str)

            # Si hay columnas adicionales, localizar los índices de las columnas inválidas
            extra_col_indices = []
            if extra_columns_flag:
                extra_cols = [c.strip() for c in columnas_adicionales_str.split(',') if c.strip()]
                # Buscar índices de las columnas adicionales por nombre en la cabecera
                for cell in ws[1]:
                    if cell.value in extra_cols:
                        extra_col_indices.append(cell.column)

            # Mapeo de errores sin campo a sus columnas específicas (de localizaciones)
            error_campo_map_loc = {
                "localizacion_duplicada": "subposicion",
                "subposicion_duplicada_excel": "subposicion",
                "caja_inconsistente": "caja",
                "rack_inconsistente": "rack",
                "posicion_rack_ocupada": "posicion_rack_estante",
                "posicion_caja_ocupada": "posicion_caja_rack"
            }
            # Recorrer filas con errores
            for fila_numero, info in errores.items():
                has_error = bool(info.get("bloqueantes"))
                if not has_error:
                    continue

                # Colorear celdas específicas y construir mensajes
                mensajes = []
                for err in info.get("bloqueantes", []):
                    if ":" in err:
                        tipo, campo = err.split(":")
                        msg = f"(Error) {MENSAJES_ERROR[tipo]}"
                        if msg not in mensajes:
                            mensajes.append(msg)
                        if campo in columnas_excel:
                            col = columnas_excel[campo]
                            celda = ws.cell(row=int(fila_numero), column=col)
                            celda.fill = FILL_ERROR_CELL
                    else:
                        campo = error_campo_map_loc.get(err)
                        msg = f"(Error) {MENSAJES_ERROR[err]}"
                        if msg not in mensajes:
                            mensajes.append(msg)
                        if campo and campo in columnas_excel:
                            col_err = columnas_excel[campo]
                            ws.cell(row=int(fila_numero), column=col_err).fill = FILL_ERROR_CELL

                celda_errores = ws.cell(row=int(fila_numero), column=col_errores)
                celda_errores.value = "\n".join(mensajes)
            
            # DESPUÉS PINTAR LAS COLUMNAS EXTRAS (sobrescribe el color de fila con rojo fuerte)
            if extra_col_indices:
                for col_idx in extra_col_indices:
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.fill = FILL_ERROR_COL
            
            output = io.BytesIO()    
            wb.save(output)
            wb.close()
            response = HttpResponse(output.getvalue(), content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="listado_errores.xlsx"'
            return response        
    else:
        form = UploadExcel(request)     
    return render(request, 'localizacion_nueva.html', {'form': form}) 

def detalles_congelador(request, nombre_congelador):
    # Vista para ver los detalles de un congelador específico
    freezer= Congelador.objects.filter(congelador=nombre_congelador)
    template=loader.get_template('detalles_congelador.html')
    return HttpResponse(template.render({'congelador':freezer[0]},request))
@permission_required('muestras.can_add_localizaciones_web')
def editar_congelador(request,nombre_congelador):
    # Vista para editar un congelador existente, requiere permiso para añadir localizaciones
    congelador = Congelador.objects.filter(congelador=nombre_congelador)
    congelador=congelador[0]
    nombre_anterior = congelador.congelador  # Capturar ANTES de is_valid()
    if request.method == 'POST':
        form = Congeladorform(request.POST, request.FILES, instance=congelador)
        if form.is_valid():
            nombre_nuevo = form.cleaned_data.get('congelador')

            with connection.cursor() as cursor:
                cursor.execute("SET FOREIGN_KEY_CHECKS=0")

            try:
                with transaction.atomic():
                    if nombre_nuevo != nombre_anterior:
                        with connection.cursor() as cursor:
                            # Actualizar FK en estantes (to_field='congelador')
                            cursor.execute(
                                "UPDATE muestras_estante SET congelador_id = %s WHERE congelador_id = %s",
                                [nombre_nuevo, nombre_anterior]
                            )
                            # Actualizar campo congelador en localizaciones (historial)
                            cursor.execute(
                                "UPDATE muestras_localizacion SET congelador = %s WHERE congelador = %s",
                                [nombre_nuevo, nombre_anterior]
                            )
                    form.save()
            finally:
                with connection.cursor() as cursor:
                    cursor.execute("SET FOREIGN_KEY_CHECKS=1")

            return redirect('detalles_congelador', nombre_congelador = form.instance.congelador)
    else:
        form = Congeladorform(instance=congelador)
    return render(request, 'editar_congelador.html', {'form': form, 'congelador': congelador})

@permission_required('muestras.can_delete_localizaciones_web')
def eliminar_localizacion(request):
    # Vista para eliminar localizaciones, requiere permiso para eliminar localizaciones
    posiciones_ocupadas = []

    # Comprobar congeladores seleccionados: si cualquiera tiene subposiciones ocupadas, bloquear
    congelador_ids = request.POST.getlist('congelador')
    if congelador_ids:
        for congelador_id in congelador_ids:
            try:
                cong = Congelador.objects.get(id=congelador_id)
            except Congelador.DoesNotExist:
                continue
            if Subposicion.objects.filter(caja__rack__estante__congelador=cong, vacia=False).exists():
                posiciones_ocupadas.append(f"Congelador {cong.congelador}")

    # Comprobar estantes seleccionados
    estante_ids = request.POST.getlist('estante')
    if estante_ids:
        for estante_id in estante_ids:
            try:
                est = Estante.objects.get(id=estante_id)
            except Estante.DoesNotExist:
                continue
            if Subposicion.objects.filter(caja__rack__estante=est, vacia=False).exists():
                posiciones_ocupadas.append(f"Estante {est.numero}")

    # Verificar racks seleccionados
    rack_ids = request.POST.getlist('rack')
    if rack_ids:
        for rack_id in rack_ids:
            try:
                rack = Rack.objects.get(id=rack_id)
            except Rack.DoesNotExist:
                continue
            if Subposicion.objects.filter(caja__rack=rack, vacia=False).exists():
                posiciones_ocupadas.append(f"Rack {rack.numero}")

    # Verificar cajas seleccionadas
    caja_ids = request.POST.getlist('caja')
    if caja_ids:
        for caja_id in caja_ids:
            try:
                caja = Caja.objects.get(id=caja_id)
            except Caja.DoesNotExist:
                continue
            if Subposicion.objects.filter(caja=caja, vacia=False).exists():
                posiciones_ocupadas.append(f"Caja {caja.numero}")

    # Verificar subposiciones seleccionadas
    subposicion_ids = request.POST.getlist('subposicion')
    if subposicion_ids:
        for subposicion_id in subposicion_ids:
            try:
                subposicion = Subposicion.objects.get(id=subposicion_id)
            except Subposicion.DoesNotExist:
                continue
            if not subposicion.vacia:
                posiciones_ocupadas.append(f"Subposición {subposicion.numero}")

    # Si hay posiciones ocupadas, mostrar error y no eliminar
    if posiciones_ocupadas:
        mensaje = f"No se pueden eliminar las siguientes posiciones porque están ocupadas: {', '.join(posiciones_ocupadas[:8])}"
        if len(posiciones_ocupadas) > 8:
            mensaje += f" y {len(posiciones_ocupadas) - 8} más."
        messages.error(request, mensaje)
        return redirect('localizaciones_todas')

    # Si no hay posiciones ocupadas, proceder con la eliminación (de abajo hacia arriba)
    if subposicion_ids:
        Subposicion.objects.filter(id__in=subposicion_ids).delete()

    if caja_ids:
        Caja.objects.filter(id__in=caja_ids).delete()

    if rack_ids:
        Rack.objects.filter(id__in=rack_ids).delete()

    if estante_ids:
        Estante.objects.filter(id__in=estante_ids).delete()

    if congelador_ids:
        Congelador.objects.filter(id__in=congelador_ids).delete()

    messages.success(request, 'Posiciones eliminadas correctamente.')
    return redirect('localizaciones_todas')

    messages.success(request, 'Posiciones eliminadas correctamente.')
    return redirect('localizaciones_todas')


def historial_localizaciones_muestra(request,muestra_id):
    # Vista para ver el historial de localizaciones de una muestra específica
    muestra = Muestra.objects.get(id=muestra_id)
    historiales = historial_localizaciones.objects.filter(muestra=muestra).order_by('-fecha_asignacion')
    if muestra.estado_actual=='Destruida':
        estado_destruccion = registro_destruido.objects.filter(muestra=muestra).first()
    else:
        estado_destruccion = None
    template = loader.get_template('historial_localizaciones.html')
    return HttpResponse(template.render({'historiales':historiales, 'muestra':muestra, 'estado_destruccion':estado_destruccion},request))

# Vistas relacionadas con el modelo estudio
@login_required
@permission_required('muestras.can_view_estudios_web')
def estudios_todos(request):
    # Vista para ver todos los estudios, los investigadores solo ven los suyos asociados
    if request.user.groups.filter(name='Investigadores'):
        queryset = Estudio.objects.filter(investigadores_asociados=request.user).annotate(num_muestras=Count('muestra'))
    else:
        queryset = Estudio.objects.all().annotate(num_muestras=Count('muestra'))

    # Búsqueda general (server-side)
    busqueda_general = request.GET.get('busqueda', '').strip()
    if busqueda_general:
        q_busqueda = Q()
        q_busqueda |= Q(referencia_estudio__icontains=busqueda_general)
        q_busqueda |= Q(nombre_estudio__icontains=busqueda_general)
        q_busqueda |= Q(descripcion_estudio__icontains=busqueda_general)
        q_busqueda |= Q(investigador_principal__icontains=busqueda_general)
        queryset = queryset.filter(q_busqueda)

    # Paginación
    contador_total = queryset.count()
    items_por_pagina = request.GET.get('items_por_pagina', 25)
    if str(items_por_pagina) == 'todas':
        items_por_pagina = 'todas'
        paginator = Paginator(queryset, max(contador_total, 1))
    else:
        try:
            items_por_pagina = int(items_por_pagina)
            if items_por_pagina not in [10, 25, 50, 100]:
                items_por_pagina = 25
        except Exception:
            items_por_pagina = 25
        paginator = Paginator(queryset, items_por_pagina)
    numero_pagina = request.GET.get('page', 1)
    try:
        estudios_page = paginator.page(numero_pagina)
    except PageNotAnInteger:
        estudios_page = paginator.page(1)
    except EmptyPage:
        estudios_page = paginator.page(paginator.num_pages)

    template = loader.get_template('estudios_todos.html')
    context = {
        'estudios': estudios_page.object_list,
        'paginator': paginator,
        'muestras_page': estudios_page,  # mantiene compatibilidad con la plantilla existente
        'contador_muestras': contador_total,
        'items_por_pagina': items_por_pagina,
        'busqueda': busqueda_general,
        'request': request,
    }
    return HttpResponse(template.render(context, request))
@permission_required('muestras.can_add_estudios_web')
def nuevo_estudio(request):
    # Vista para crear un nuevo estudio
    if request.method == 'POST':
        form = EstudioForm(request.POST)
        if form.is_valid():
            # Comprobar que la referencia no existe ya en la base de datos
            referencia = form.cleaned_data.get('referencia_estudio')
            if referencia:
                referencia_norm = str(referencia).strip()
                if Estudio.objects.filter(referencia_estudio__iexact=referencia_norm).exists():
                    form.add_error('referencia_estudio', 'La referencia indicada ya existe en la base de datos.')
                    template = loader.get_template('nuevo_estudio.html')
                    return HttpResponse(template.render({'form': form}, request))

            # Validar que fecha_inicio <= fecha_fin si ambas están presentes
            fecha_inicio = form.cleaned_data.get('fecha_inicio_estudio')
            fecha_fin = form.cleaned_data.get('fecha_fin_estudio')
            if fecha_inicio and fecha_fin and fecha_fin < fecha_inicio:
                form.add_error('fecha_fin_estudio', 'La fecha de fin debe ser igual o posterior a la fecha de inicio.')
                template = loader.get_template('nuevo_estudio.html')
                return HttpResponse(template.render({'form': form}, request))

            form.save()
            messages.success(request,'Estudio añadido correctamente')
            return redirect('estudios_todos')
    else:
        form = EstudioForm()
    template = loader.get_template('nuevo_estudio.html')
    return HttpResponse(template.render({'form':form},request))
@permission_required('muestras.can_add_estudios_web')
def excel_estudios(request):
    # Mostrar confirmación después de validación con progreso
    if request.GET.get('mostrar_confirmacion') and 'confirmacion_pendiente' in request.session:
        datos_conf = request.session.pop('confirmacion_pendiente')
        for msg in datos_conf.get('mensajes', []):
            getattr(messages, msg['level'])(request, msg['text'])
        return render(request, datos_conf['template'], datos_conf.get('context', {}))
    # Vista para subir estudios desde un archivo Excel, requiere permiso para añadir estudios
    if request.method=="POST":
        form = UploadExcel(request.POST, request.FILES)
        # Si el usuario confirma, se guardan los estudios en la base de datos
        if 'confirmar' in request.POST:
            filas_validas = request.session.get('filas_validas',[])
            total = len(filas_validas)

            def gen_confirmar_estudios():
                yield _progress_page_start('Importando estudios', total)
                try:
                    with transaction.atomic():
                        for i, datos in enumerate(filas_validas):
                            fecha_inicio_estudio = date.fromisoformat(datos["fecha_inicio_estudio"]) if datos["fecha_inicio_estudio"] else None
                            fecha_fin_estudio = date.fromisoformat(datos["fecha_fin_estudio"]) if datos["fecha_fin_estudio"] else None
                            Estudio.objects.create(
                                referencia_estudio=datos['referencia_estudio'],
                                nombre_estudio=datos['nombre_estudio'],
                                descripcion_estudio=datos['descripcion_estudio'],
                                fecha_inicio_estudio=fecha_inicio_estudio,
                                fecha_fin_estudio=fecha_fin_estudio,
                                investigador_principal=datos['investigador_principal']
                            )
                            if _should_update(i, total):
                                yield _progress_update(i + 1, total)
                    yield _progress_done('/estudios/', 'Estudios importados correctamente')
                except Exception as e:
                    yield _progress_error(str(e))

            return StreamingHttpResponse(gen_confirmar_estudios(), content_type='text/html')
        # Si el usuario cancela, no se hace nada
        elif 'cancelar' in request.POST:
            messages.error(request,'Los estudios no se han añadido')
            return redirect('estudios_todos')
        # Si se sube un archivo excel, se procesa y valida
        elif 'excel_file' in request.FILES:
            if form.is_valid():
                # Leer excel y preparar columnas
                excel_file = request.FILES['excel_file']
                if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
                    return render(request, 'upload_excel_estudios.html', {'form': form, 'error': '❌ Error de formato: El archivo debe ser un Excel (.xlsx o .xls).'})
                excel_bytes = excel_file.read()
                request.session['excel_file_name'] = excel_file.name
                request.session['excel_file_base64']= base64.b64encode(excel_bytes).decode()
                excel_stream = io.BytesIO(excel_bytes)
                df = pd.read_excel(excel_stream)
                if df.empty:
                    return render(request, 'upload_excel_estudios.html', {'form': form, 'error': '❌ Error de formato: El archivo Excel está vacío o no contiene filas de datos.'})
                rename_columns = {
                    'Referencia del estudio': 'referencia_estudio', 
                    'Nombre del estudio': 'nombre_estudio',
                    'Descripción': 'descripcion_estudio',
                    'Fecha de inicio': 'fecha_inicio_estudio',
                    'Fecha de fin': 'fecha_fin_estudio',
                    'Investigador principal': 'investigador_principal',
                }
                # Validar columnas
                columnas_esperadas = set(rename_columns.keys())
                columnas_existentes = set(df.columns)
                columnas_faltantes = columnas_esperadas - columnas_existentes
                if columnas_faltantes:
                    columnas_str = ', '.join(sorted(columnas_faltantes))
                    return render(request, 'upload_excel_estudios.html', {'form': form, 'error': f'❌ Error de formato: El archivo Excel no contiene las siguientes columnas esperadas: {columnas_str}'})
                columnas_adicionales = columnas_existentes - columnas_esperadas
                if columnas_adicionales:
                    columnas_adicionales_str = ', '.join(sorted(columnas_adicionales))
                    request.session['columnas_adicionales'] = columnas_adicionales_str
                df.rename(columns=rename_columns, inplace=True)
                 # Funciones para normalizar las columnas del excel
                def norm(value):
                    if value is None or pd.isna(value):
                        return None

                    if isinstance(value, str):
                        value = value.strip()
                        return value if value != "" else None

                    return value
                '''
                def norm_code(value):
                    if value is None or pd.isna(value):
                        return None

                    if isinstance(value, float) and value.is_integer():
                        return str(int(value))

                    return str(value).strip()
                def convertir_fecha(value):
                        if pd.isna(value) or value is None:
                            return None
                        fecha = pd.to_datetime(value)
                        return fecha.date()
                '''
                # Crear estructuras previas del excel
                total_filas = len(df)

                def gen_validar_estudios():
                    filas_validas = []
                    errores = {}
                    nombre_estudios_excel = set()
                    numero_registros = 0
                    cache_inner = {
                        'estudios_existentes_lower': set(n.lower() for n in Estudio.objects.values_list('nombre_estudio', flat=True).distinct() if n),
                        'referencias_existentes_lower': set(str(x).strip().lower() for x in Estudio.objects.values_list('referencia_estudio', flat=True).distinct() if x is not None)
                    }
                    referencias_excel = set()

                    yield _progress_page_start('Validando estudios', total_filas)

                    for idx, row in df.iterrows():
                        # Recorrer el df para detectar errores y normalizar
                        numero_registros += 1
                        fila = idx + 2 
                        errores[fila]={"advertencias":[], "bloqueantes":[]}
                        datos = {
                            "nombre_estudio":norm(row['nombre_estudio']),
                            'referencia_estudio': norm(row['referencia_estudio']),
                            'descripcion_estudio': norm(row['descripcion_estudio']),
                            'fecha_inicio_estudio':norm(row['fecha_inicio_estudio']),
                            'fecha_fin_estudio': norm(row['fecha_fin_estudio']),
                            'investigador_principal': norm(row['investigador_principal'])
                        }
                        # Detectar campos vacios
                        optativos = ["referencia_estudio", "descripcion_estudio", "fecha_inicio_estudio", "fecha_fin_estudio", "investigador_principal"]
                        for campo in optativos:
                            if not datos.get(campo):
                                errores[fila]["advertencias"].append(f"campo_optativo_vacio:{campo}")
                        obligatorios = ["nombre_estudio"]
                        for campo in obligatorios:
                            if not datos.get(campo):
                                errores[fila]["bloqueantes"].append(f"campo_obligatorio_vacio:{campo}")
                        
                        # Validar que ningún campo contenga el carácter punto y coma (;)
                        campos_a_validar = ['referencia_estudio', 'nombre_estudio', 'descripcion_estudio', 'investigador_principal']
                        for campo in campos_a_validar:
                            valor = datos.get(campo)
                            if valor and isinstance(valor, str) and ';' in valor:
                                errores[fila]["bloqueantes"].append(f"caracter_invalido_semicolon:{campo}")
                        
                        # Detectar formato de fecha incorrecto (advertencia)
                        for campo in ['fecha_inicio_estudio', 'fecha_fin_estudio']:
                            if datos[campo] != None:
                                try:
                                    # Si es un Timestamp de pandas o datetime, usar directamente
                                    if isinstance(datos[campo], (pd.Timestamp, type(pd.NaT))):
                                        if pd.isna(datos[campo]):
                                            datos[campo] = None
                                        else:
                                            # Convertir Timestamp/datetime a ISO string
                                            datos[campo] = datos[campo].date().isoformat()
                                    elif isinstance(datos[campo], date):
                                        # Si es un objeto date, convertir directamente a ISO
                                        datos[campo] = datos[campo].isoformat()
                                    else:
                                        # Si es string, parsear con formato DD-MM-AAAA
                                        fecha_str = str(datos[campo]).strip()
                                        partes = fecha_str.split('-')
                                        if len(partes) == 3 and all(p.isdigit() for p in partes):
                                            fecha = pd.to_datetime(fecha_str, format='%d-%m-%Y')
                                            datos[campo] = fecha.date().isoformat()
                                        else:
                                            errores[fila]["bloqueantes"].append(f"fecha_invalida:{campo}")
                                            datos[campo] = None
                                except Exception:
                                    errores[fila]["bloqueantes"].append(f"fecha_invalida:{campo}")
                                    datos[campo] = None

                        # Validar que fecha_fin >= fecha_inicio si ambas están informadas
                        fecha_inicio = datos.get('fecha_inicio_estudio')
                        fecha_fin = datos.get('fecha_fin_estudio')
                        if fecha_inicio and fecha_fin:
                            # Ambas fechas están informadas y son válidas
                            if fecha_fin < fecha_inicio:
                                errores[fila]["bloqueantes"].append("fecha_fin_menor_que_inicio")

                        # Detectar si el estudio ya existe
                        nombre_estudio = datos['nombre_estudio']
                        if nombre_estudio:
                            # Normalizar a string en caso de que sea numérico
                            nombre_estudio_str = str(nombre_estudio).strip()
                            nombre_estudio_lower = nombre_estudio_str.lower()
                            if nombre_estudio_lower in cache_inner['estudios_existentes_lower']:
                                errores[fila]["bloqueantes"].append(f"estudio_existente")
                            if nombre_estudio_lower in nombre_estudios_excel:
                                errores[fila]["bloqueantes"].append("estudio_duplicado_excel")
                            else:
                                nombre_estudios_excel.add(nombre_estudio_lower)
                        else:
                            nombre_estudio_lower = ''
                        # Validar referencia_estudio: si existe, no puede coincidir con otras en DB ni duplicarse en el Excel
                        referencia = datos.get('referencia_estudio')
                        if referencia:
                            # Normalizar a string antes de usar lower() para admitir valores numéricos en Excel
                            ref_str = str(referencia).strip()
                            if ref_str:
                                ref_lower = ref_str.lower()
                                if ref_lower in cache_inner['referencias_existentes_lower']:
                                    errores[fila]["bloqueantes"].append("referencia_existente")
                                if ref_lower in referencias_excel:
                                    errores[fila]["bloqueantes"].append("referencia_duplicada_excel")
                                else:
                                    referencias_excel.add(ref_lower)
                       
                        # Registrar filas validas
                        if not errores[fila]["bloqueantes"]:
                            filas_validas.append(datos)

                        if _should_update(idx, total_filas):
                            yield _progress_update(idx + 1, total_filas)

                    request.session['filas_validas'] = filas_validas
                    request.session['errores'] = errores

                    # Obtener configuración de mensajes para estudios
                    msg_config = get_upload_messages('estudios')

                    # Contar errores
                    numero_errores_bloqueantes = 0
                    numero_errores_advertencia = 0
                    for fila in errores:
                        if errores[fila]['bloqueantes']:
                            numero_errores_bloqueantes += 1
                        if errores[fila]["advertencias"]:
                            numero_errores_advertencia += 1

                    mensajes = []

                    # Mensaje inicial
                    mensajes.append({'level': 'info', 'text': f'{msg_config["titulo_inicial"]} {numero_registros} registros.'})

                    # Generar mensajes según el estado
                    if numero_errores_advertencia > 0:
                        msg = msg_config['con_advertencias'].format(count=numero_errores_advertencia)
                        mensajes.append({'level': 'warning', 'text': msg})
                    if numero_errores_bloqueantes > 0:
                        msg = msg_config['con_bloqueantes'].format(count=numero_errores_bloqueantes)
                        mensajes.append({'level': 'error', 'text': msg})
                    if numero_errores_bloqueantes == 0 and numero_errores_advertencia == 0:
                        mensajes.append({'level': 'success', 'text': msg_config['sin_errores']})

                    # Manejar columnas extras
                    columnas_extras_str = request.session.get('columnas_adicionales', '')
                    tiene_columnas_extras = bool(columnas_extras_str)
                    numero_columnas_extras = len(columnas_extras_str.split(', ')) if columnas_extras_str else 0
                    if tiene_columnas_extras:
                        msg = msg_config['columnas_extras'].format(count=numero_columnas_extras, detalles=columnas_extras_str)
                        mensajes.append({'level': 'warning', 'text': msg})

                    request.session['confirmacion_pendiente'] = {
                        'template': 'confirmacion_upload_estudios.html',
                        'context': {
                            'numero_errores_bloqueantes': numero_errores_bloqueantes,
                            'numero_errores_advertencia': numero_errores_advertencia,
                            'tiene_columnas_extras': tiene_columnas_extras,
                            'numero_columnas_extras': numero_columnas_extras,
                            'columnas_extras_str': columnas_extras_str
                        },
                        'mensajes': mensajes
                    }
                    request.session.save()
                    yield _progress_done(request.path + '?mostrar_confirmacion=1', 'Validación completada')

                return StreamingHttpResponse(gen_validar_estudios(), content_type='text/html')
            
        # Si se solicita un excel de errores, este se rellena en base a los errores detectados durante la validación
        elif 'excel_errores' in request.POST:
                    errores = request.session.get('errores',[])
                    excel_bytes = base64.b64decode(request.session.get('excel_file_base64'))
                    excel_file = io.BytesIO(excel_bytes)
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    # Definir los estilos para pintar el excel usando configuración centralizada
                    colors = get_excel_colors()
                    FILL_ERROR_CELL = PatternFill("solid", fgColor=colors['error_cell'])
                    FILL_WARN_CELL  = PatternFill("solid", fgColor=colors['warning_cell'])
                    FILL_EXTRA_COL  = PatternFill("solid", fgColor=colors['extra_column'])
                    # Diccionario de mensajes
                    MENSAJES_ERROR = {
                        "campo_obligatorio_vacio": "Campo obligatorio vacío",
                        "fecha_invalida": "Fecha inválida (Formato correcto: DD-MM-AAAA)",
                        "fecha_fin_menor_que_inicio": "Fecha de fin anterior a fecha de inicio",
                        "estudio_existente": "El estudio ya existe en la base de datos",
                        "estudio_duplicado_excel": "Estudio duplicado dentro del Excel",
                        "referencia_existente": "Referencia ya existe en la base de datos",
                        "referencia_duplicada_excel": "Referencia duplicada dentro del Excel",
                        "campo_optativo_vacio": "Campo opcional vacío",
                        "caracter_invalido_semicolon": "El carácter ';' no está permitido en este campo"
                    }
                    # Diccionario de columnas del excel
                    columnas_excel = {}
                    rename_columns = {
                    'Referencia del estudio': 'referencia_estudio', 
                    'Nombre del estudio': 'nombre_estudio',
                    'Descripción': 'descripcion_estudio',
                    'Fecha de inicio': 'fecha_inicio_estudio',
                    'Fecha de fin': 'fecha_fin_estudio',
                    'Investigador principal': 'investigador_principal',
                    }
                    for cell in ws[1]:
                        columnas_excel[rename_columns.get(cell.value, cell.value)] = cell.column
                    # Añadir la columna de errores
                    col_errores = ws.max_column + 1
                    ws.cell(row=1, column=col_errores, value="Errores")
                    # Mapeo de errores sin campo a sus columnas específicas (estudios)
                    error_campo_map_study = {
                        "estudio_existente": "nombre_estudio",
                        "estudio_duplicado_excel": "nombre_estudio",
                        "referencia_existente": "referencia_estudio",
                        "referencia_duplicada_excel": "referencia_estudio"
                    }
                    # Recorrer filas con errores 
                    for fila, info in errores.items():
                        has_error = bool(info.get("bloqueantes", []))
                        has_warn = bool(info.get("advertencias", []))
                        if not has_error and not has_warn:
                            continue

                        # Colorear celdas específicas y construir mensajes
                        mensajes = []
                        for err in info.get("bloqueantes", []):
                            if ":" in err:
                                tipo, campo = err.split(":")
                                msg = f"(Error) {MENSAJES_ERROR[tipo]}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                                col = columnas_excel[campo]
                                celda = ws.cell(row=int(fila), column=col)
                                celda.fill = FILL_ERROR_CELL
                            else:
                                campo = error_campo_map_study.get(err)
                                msg = f"(Error) {MENSAJES_ERROR[err]}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                                if campo and campo in columnas_excel:
                                    col_err = columnas_excel[campo]
                                    ws.cell(row=int(fila), column=col_err).fill = FILL_ERROR_CELL
                        for warn in info.get("advertencias", []): 
                            if ":" in warn:
                                tipo, campo = warn.split(":", 1)
                                mensaje_warn = MENSAJES_ERROR.get(tipo, tipo)
                                msg = f"(Advertencia) {mensaje_warn}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                                if campo in columnas_excel:
                                    col = columnas_excel[campo]
                                    celda = ws.cell(row=int(fila), column=col)
                                    celda.fill = FILL_WARN_CELL
                            else:
                                mensaje_warn = MENSAJES_ERROR.get(warn, warn)
                                msg = f"(Advertencia) {mensaje_warn}"
                                if msg not in mensajes:
                                    mensajes.append(msg)
                        ws.cell(row=int(fila), column=col_errores, value="\n".join(mensajes))

                    # Pintar columnas extras con color de columna extra
                    expected_renamed = set(rename_columns.values())
                    for col_name, col_num in columnas_excel.items():
                        if col_name not in expected_renamed:
                            # Pintar el encabezado
                            header_cell = ws.cell(row=1, column=col_num)
                            header_cell.fill = FILL_EXTRA_COL
                            # Pintar todas las celdas de datos en la columna
                            for row in range(2, ws.max_row + 1):
                                ws.cell(row=row, column=col_num).fill = FILL_EXTRA_COL

                    output = io.BytesIO()    
                    wb.save(output)
                    wb.close()
                    response = HttpResponse(output.getvalue(),content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="listado_errores.xlsx"'
                    return response      
    else: 
        form= UploadExcel()
    return render(request, 'upload_excel_estudios.html', {'form': form}) 
@permission_required('muestras.can_change_estudios_web')
def editar_estudio(request, id_estudio):
    # Vista para editar un estudio existente
    estudio = Estudio.objects.get(id=id_estudio)
    nombre_anterior = estudio.nombre_estudio  # Capturar ANTES de is_valid()
    if request.method == 'POST':
        form = EstudioForm(request.POST, instance=estudio)
        if form.is_valid():
            # Validar que no existan duplicados de nombre o referencia (excluyendo el estudio actual)
            nombre = form.cleaned_data.get('nombre_estudio')
            referencia = form.cleaned_data.get('referencia_estudio')
            duplicate = False
            if nombre:
                nombre_norm = str(nombre).strip()
                qs_nombre = Estudio.objects.filter(nombre_estudio__iexact=nombre_norm).exclude(id=estudio.id)
                if qs_nombre.exists():
                    form.add_error('nombre_estudio', 'Ya existe otro estudio con ese nombre.')
                    duplicate = True
            if referencia:
                referencia_norm = str(referencia).strip()
                qs_ref = Estudio.objects.filter(referencia_estudio__iexact=referencia_norm).exclude(id=estudio.id)
                if qs_ref.exists():
                    form.add_error('referencia_estudio', 'Ya existe otro estudio con esa referencia.')
                    duplicate = True

            # Validar que fecha_inicio <= fecha_fin si ambas están presentes
            fecha_inicio = form.cleaned_data.get('fecha_inicio_estudio')
            fecha_fin = form.cleaned_data.get('fecha_fin_estudio')
            if fecha_inicio and fecha_fin and fecha_fin < fecha_inicio:
                form.add_error('fecha_fin_estudio', 'La fecha de fin debe ser igual o posterior a la fecha de inicio.')
                duplicate = True

            if duplicate:
                template = loader.get_template('editar_estudio.html')
                return HttpResponse(template.render({'form': form, 'estudio': estudio}, request))

            nombre_nuevo = form.cleaned_data.get('nombre_estudio')

            with connection.cursor() as cursor:
                cursor.execute("SET FOREIGN_KEY_CHECKS=0")

            try:
                with transaction.atomic():
                    # Si el nombre cambió, actualizar las referencias en muestras y localizaciones
                    if nombre_nuevo != nombre_anterior:
                        with connection.cursor() as cursor:
                            cursor.execute(
                                "UPDATE muestras_muestra SET estudio_id = %s WHERE estudio_id = %s",
                                [nombre_nuevo, nombre_anterior]
                            )

                    # Preservar fechas si el usuario las dejó vacías pero había valores anteriores
                    estudio_guardado = form.save(commit=False)
                    if not form.cleaned_data.get('fecha_inicio_estudio') and estudio.fecha_inicio_estudio:
                        estudio_guardado.fecha_inicio_estudio = estudio.fecha_inicio_estudio
                    if not form.cleaned_data.get('fecha_fin_estudio') and estudio.fecha_fin_estudio:
                        estudio_guardado.fecha_fin_estudio = estudio.fecha_fin_estudio
                    estudio_guardado.save()
            finally:
                with connection.cursor() as cursor:
                    cursor.execute("SET FOREIGN_KEY_CHECKS=1")

            messages.info(request,'El estudio se ha modificado correctamente')
            return redirect('estudios_todos')
    else:
        form = EstudioForm(instance=estudio)
    return render(request, 'editar_estudio.html', {'form': form, 'estudio': estudio})
@permission_required('muestras.can_delete_estudios_web')
def eliminar_estudio(request, id_estudio):
    # Vista para eliminar un estudio existente
    estudio = get_object_or_404(Estudio,id=id_estudio)
    if Muestra.objects.filter(estudio=estudio).exists():
        messages.error(request, mark_safe(f'No se puede eliminar el estudio "{estudio.nombre_estudio}" porque tiene muestras asociadas. Desasocia las muestras primero.'))
        return redirect('estudios_todos')
    estudio.delete()
    messages.success(request,'Estudio eliminado correctamente')
    return redirect('estudios_todos')
@permission_required('muestras.can_change_estudios_web')
def seleccionar_estudio(request):
    # Vista para seleccionar un estudio al que añadir muestras
    estudios = Estudio.objects.all()
    template = loader.get_template('seleccionar_estudio.html')
    return HttpResponse(template.render({'estudios':estudios},request))
@permission_required('muestras.can_change_estudios_web')
def añadir_muestras_estudio(request):
    # Vista para añadir muestras a un estudio seleccionado
    if request.method == 'POST':
        # Obtener las muestras de la sesión
        muestras = request.session.get('muestras_estudio', [])
        muestras=Muestra.objects.filter(id__in=muestras)
        # Desasociar muestras de sus estudios si se selecciona esa opción
        if len(request.POST.getlist('desasociar')) ==1:
            for muestra in muestras:
                if muestra.estado_actual != 'Destruida':
                    muestra.estudio = None
                    muestra.save()
                    historial = historial_estudios.objects.create(
                            muestra = muestra,
                            estudio = None,
                            fecha_asignacion = timezone.now(),
                            usuario_asignacion = request.user
                        )
                    historial.save()

            return redirect('muestras_todas')
        # Obtener los estudios seleccionados y asociar las muestras
        ids_estudios = request.POST.getlist('estudio_nombre')
        for study in ids_estudios:
            studio = Estudio.objects.get(nombre_estudio=study)
            for muestra in muestras:
                if muestra.estado_actual != 'Destruida':
                    muestra.estudio = studio
                    muestra.save()
                    # Crear entrada en el historial de estudios si la muestra no estaba ya asociada a ese estudio
                    if historial_estudios.objects.filter(muestra=muestra,estudio=studio).exists():
                        pass
                    else:   
                        historial = historial_estudios.objects.create(
                            muestra = muestra,
                            estudio = studio,
                            fecha_asignacion = timezone.now(),
                            usuario_asignacion = request.user
                        )
                        historial.save()
            messages.success(request,'Muestras añadidas correctamente a los estudios')
        if 'muestras_estudio' in request.session:
            del request.session['muestras_estudio']
        return redirect('muestras_todas')
    return redirect('muestras_todas')

def historial_estudios_muestra(request,muestra_id):
    # Vista para ver el historial de estudios de una muestra específica
    muestra = Muestra.objects.get(id=muestra_id)
    historiales = historial_estudios.objects.filter(muestra=muestra).order_by('-fecha_asignacion')
    template = loader.get_template('historial_estudios.html')
    return HttpResponse(template.render({'historiales':historiales, 'muestra':muestra},request))
@permission_required('muestras.can_view_estudios_web')
def repositorio_estudio(request, id_estudio):
    # Vista para ver el repositorio de documentos asociado a un estudio
    estudio = Estudio.objects.get(id=id_estudio)
    documentos = Documento.objects.filter(estudio = estudio, eliminado= False)
    request.session['id'] = id_estudio
    usuarios = User.objects.all()
    # Filtrado opcional por usuario (soporta múltiples valores separados por ;)
    usuario = request.GET.get('usuario')
    if usuario:
        usuarios_list = [u.strip() for u in usuario.split(';') if u.strip()]
        if usuarios_list:
            q_filters = Q()
            for user in usuarios_list:
                q_filters |= Q(usuario_subida__username=user)
            documentos = documentos.filter(q_filters)
    # Filtrado opcional por categoría (soporta múltiples valores separados por ;)
    categoria = request.GET.get('categoria')
    if categoria:
        categorias_list = [c.strip() for c in categoria.split(';') if c.strip()]
        if categorias_list:
            q_filters = Q()
            for cat in categorias_list:
                q_filters |= Q(categoria__icontains=cat)
            documentos = documentos.filter(q_filters)
    # Búsqueda general de documentos (server-side)
    busqueda_general = request.GET.get('busqueda', '').strip()
    if busqueda_general:
        q_busqueda = Q()
        q_busqueda |= Q(categoria__icontains=busqueda_general)
        q_busqueda |= Q(descripcion__icontains=busqueda_general)
        q_busqueda |= Q(usuario_subida__username__icontains=busqueda_general)
        q_busqueda |= Q(archivo__icontains=busqueda_general)
        documentos = documentos.filter(q_busqueda)

    for doc in documentos:    
        if request.GET.get(f'{doc.id}'):
            eliminar_documento(request, doc.id)
    # Paginación de documentos
    contador_total = documentos.count()
    items_por_pagina = request.GET.get('items_por_pagina', 25)
    if str(items_por_pagina) == 'todas':
        items_por_pagina = 'todas'
        paginator = Paginator(documentos, max(contador_total, 1))
    else:
        try:
            items_por_pagina = int(items_por_pagina)
            if items_por_pagina not in [10, 25, 50, 100]:
                items_por_pagina = 25
        except Exception:
            items_por_pagina = 25
        paginator = Paginator(documentos, items_por_pagina)
    numero_pagina = request.GET.get('page', 1)
    try:
        documentos_page = paginator.page(numero_pagina)
    except PageNotAnInteger:
        documentos_page = paginator.page(1)
    except EmptyPage:
        documentos_page = paginator.page(paginator.num_pages)

    template = loader.get_template('repositorio_estudio.html')
    context = {
        'documentos': documentos_page.object_list,
        'paginator': paginator,
        'muestras_page': documentos_page,  # mantiene compatibilidad con la plantilla existente
        'contador_muestras': contador_total,
        'items_por_pagina': items_por_pagina,
        'busqueda': busqueda_general,
        'id': estudio.id,
        'estudio': estudio,
        'usuarios': usuarios,
        'request': request,
    }
    return HttpResponse(template.render(context, request))

def subir_documento(request, id_estudio):
    # Vista para subir un documento a un estudio específico
    estudio = Estudio.objects.get(id = id_estudio)
    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            # Definir el usuario que sube el documento y el estudio asociado
            doc.usuario_subida = request.user
            doc.estudio = estudio
            doc.save()
            return redirect('repositorio_estudio', id_estudio=estudio.id)
        else:
            messages.error(request, 'Hubo un error al subir el documento.')
    else:
        form = DocumentoForm()
    template = loader.get_template('subir_documento.html')
    return HttpResponse(template.render({'form':form, 'estudio':estudio},request))

def descargar_documento(request, id_estudio, documento_id):
    # Vista para descargar un documento del repositorio de un estudio
    doc = Documento.objects.get(pk=documento_id, eliminado=False)      
    return FileResponse(open(doc.archivo.path, 'rb'), as_attachment=True, filename=os.path.basename(doc.archivo.name))

def eliminar_documento(request):
    # Vista para eliminar documentos del repositorio de un estudio
    ids_documento = request.POST.getlist('doc_id')
    for element in ids_documento:
        try:
            doc = Documento.objects.get(pk=element, eliminado=False)
            # Obtener id_estudio antes de eliminar
            id_estudio = doc.estudio.id
            # Eliminar el archivo físico del servidor
            if doc.archivo:
                ruta_archivo = os.path.join(settings.MEDIA_ROOT, doc.archivo.name)
                if os.path.exists(ruta_archivo):
                    os.remove(ruta_archivo)
            # Eliminar el documento de la base de datos
            doc.delete()
            return redirect('repositorio_estudio', id_estudio=id_estudio)
        except:
            return redirect('repositorio_estudio', id_estudio=request.session['id'])
    return redirect('repositorio_estudio', id_estudio=request.session['id'])

# Vistas relacionadas con el envio de muestras
@permission_required('muestras.can_change_muestras_web')
def formulario_envios(request,centro):
    # Vista para mostrar el formulario de envíos de muestras a un centro específico
    muestras_envio = request.session.get('muestras_envio', [])
    centro_envio = agenda_envio.objects.get(id=centro)
    muestras = Muestra.objects.filter(id__in=muestras_envio, volumen_actual__gt=0)
    template = loader.get_template('formulario_envios.html')
    return HttpResponse(template.render({'muestras':muestras,'centro':centro_envio},request))

def registrar_envio(request,centro):
    # Vista para registrar el envío de muestras a un centro específico desde el formulario de envíos
    if request.method=='POST':
        # Obtener los datos del formulario, guardados en la sesión y registrar los envíos
        centro_envio = agenda_envio.objects.get(id=centro)
        muestras = request.session.get('muestras_envio', [])
        volumen_enviado_form = request.POST.getlist('volumen_enviado')
        concentracion_enviada_form = request.POST.getlist('concentracion_enviada')
        centro_destino_form = centro_envio.centro
        lugar_destino_form = centro_envio.lugar
        iterar = 0
        for muestra in muestras:
            instancia_muestra = Muestra.objects.get(id=muestra)
            envio = Envio.objects.create(
                muestra=instancia_muestra,
                fecha_envio=timezone.now(),
                volumen_enviado = volumen_enviado_form[iterar],
                unidad_volumen_enviado = instancia_muestra.unidad_volumen,
                concentracion_enviada = concentracion_enviada_form[iterar],
                unidad_concentracion_enviada = instancia_muestra.unidad_concentracion,
                centro_destino = centro_destino_form,
                lugar_destino=lugar_destino_form,
                usuario_envio = request.user
            )
            envio.save()
            # Actualizar el estado, la posición y el volumen de la muestra tras el envío
            if float(volumen_enviado_form[iterar]) >= instancia_muestra.volumen_actual:
                instancia_muestra.volumen_actual = 0
                instancia_muestra.concentracion_actual = 0
                instancia_muestra.estado_actual = 'ENV'
                instancia_muestra.save()
                if Localizacion.objects.filter(muestra=instancia_muestra).exists():
                    loc = Localizacion.objects.get(muestra=instancia_muestra)
                    loc.muestra = None
                    loc.save()
                if Subposicion.objects.filter(muestra=instancia_muestra).exists():
                    sub = Subposicion.objects.get(muestra=instancia_muestra)
                    sub.muestra = None
                    sub.vacia = True
                    sub.save()
            else:
                instancia_muestra.volumen_actual -= float(volumen_enviado_form[iterar])
                instancia_muestra.estado_actual = 'PENV'
                instancia_muestra.save()
            iterar += 1
        if 'muestras_envio' in request.session:
            del request.session['muestras_envio']
        return redirect('muestras_todas')
    return redirect('formulario_envios')

def upload_excel_envios(request,centro):
    # Mostrar confirmación después de validación con progreso
    if request.GET.get('mostrar_confirmacion') and 'confirmacion_pendiente' in request.session:
        datos_conf = request.session.pop('confirmacion_pendiente')
        for msg in datos_conf.get('mensajes', []):
            getattr(messages, msg['level'])(request, msg['text'])
        return render(request, datos_conf['template'], datos_conf.get('context', {}))
    # Vista para subir un archivo Excel con los datos de envío de muestras
    centro_envio = agenda_envio.objects.get(id=centro)
    if request.method=='POST':
        form = UploadExcel(request.POST, request.FILES)
        if 'confirmar' in request.POST:
            # Si el usuario confirma, se registran los envíos en la base de datos
            filas_validas = request.session.get('filas_validas',[])
            total = len(filas_validas)

            def gen_confirmar_envios():
                yield _progress_page_start('Registrando envíos', total)
                try:
                    with transaction.atomic():
                        for i, datos in enumerate(filas_validas):
                            muestra = Muestra.objects.get(nom_lab=datos['nom_lab'])
                            Envio.objects.create(
                                muestra=muestra,
                                volumen_enviado=datos['volumen_enviado'],
                                unidad_volumen_enviado=datos['unidad_volumen_enviado'],
                                concentracion_enviada=datos['concentracion_enviada'],
                                unidad_concentracion_enviada=datos['unidad_concentracion_enviada'],
                                centro_destino=datos['centro_destino'],
                                lugar_destino=datos['lugar_destino'],
                                fecha_envio=timezone.now(),
                                usuario_envio=request.user
                            )
                            if datos['volumen_enviado'] >= muestra.volumen_actual:
                                muestra.volumen_actual = 0
                                muestra.concentracion_actual = 0
                                muestra.estado_actual = 'ENV'
                                muestra.save()
                                if Subposicion.objects.filter(muestra=muestra).exists():
                                    sub = Subposicion.objects.get(muestra=muestra)
                                    sub.muestra = None
                                    sub.save()
                            else:
                                muestra.volumen_actual -= float(datos['volumen_enviado'])
                                muestra.estado_actual = 'PENV'
                                muestra.save()
                            if _should_update(i, total):
                                yield _progress_update(i + 1, total)
                    yield _progress_done('/muestras/', 'Envíos registrados correctamente')
                except Exception as e:
                    yield _progress_error(str(e))

            return StreamingHttpResponse(gen_confirmar_envios(), content_type='text/html')
        
        elif 'cancelar' in request.POST:
            # Si el usuario cancela, no se registra nada
            messages.error(request,'El envio no se ha registrado')
            return redirect('muestras_todas')
        
        elif 'descargar_excel_envio' in request.POST:
            # Si se solicita descargar el excel de envío, se genera y se rellena con los datos de las muestras a enviar, en caso de que se hayan seleccionado y no estén destruidas
            muestras = request.session.get('muestras_envio',[])
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="listado_envio.xlsx"'
            wb = openpyxl.load_workbook(os.path.join(settings.BASE_DIR, 'datos_prueba', 'globalstaticfiles', 'plantilla_envios.xlsx'))
            ws = wb.active
            row_num = 2
            for muestra in muestras:
                sample = Muestra.objects.get(id=muestra)
                if sample.estado_actual != 'Destruida':
                    ws.cell(row_num,1).value=str(sample.nom_lab)
                    if sample.volumen_actual != None:
                        ws.cell(row_num,2).value=str(sample.volumen_actual) + ' ' + str(sample.unidad_volumen)
                    if sample.concentracion_actual != None:
                        ws.cell(row_num,3).value=str(sample.concentracion_actual) + ' ' + str(sample.unidad_concentracion)
                    ws.cell(row_num,5).value=str(sample.unidad_volumen)
                    ws.cell(row_num,7).value=str(sample.unidad_concentracion)
                    ws.cell(row_num,8).value=str(centro_envio.centro)
                    ws.cell(row_num,9).value=str(centro_envio.lugar)
                    row_num +=1 
            wb.save(response)
            return response
        elif 'excel_file' in request.FILES:
            # Limpiar sesión residual de uploads anteriores
            if 'columnas_adicionales' in request.session:
                del request.session['columnas_adicionales']
            # Si se sube un archivo excel, se procesa y valida
            if form.is_valid():
                # Leer excel y preparar columnas 
                excel_file = request.FILES['excel_file']
                excel_bytes = excel_file.read()
                request.session['excel_file_name'] = excel_file.name
                request.session['excel_file_base64']= base64.b64encode(excel_bytes).decode()
                excel_stream = io.BytesIO(excel_bytes)
                df = pd.read_excel(excel_stream)
                rename_columns = {
                    'Muestra':'nom_lab',
                    'Volumen enviado':'volumen_enviado', 
                    'Volumen actual': 'volumen_actual',
                    'Concentración actual':'concentracion_actual',
                    'Concentración enviada':'concentracion_enviada',
                    'Unidad de volumen':'unidad_volumen_enviado',
                    'Unidad de concentración':'unidad_concentracion_enviada',
                    'Centro de destino':'centro_destino',
                    'Lugar de destino':'lugar_destino'
                }
                df.rename(columns=rename_columns, inplace=True)
                # Funciones para normalizar las columnas del excel
                def norm(value):
                    if value is None or pd.isna(value):
                        return None

                    if isinstance(value, str):
                        value = value.strip()
                        return value if value != "" else None

                    return value
                
                def norm_code(value):
                    if value is None or pd.isna(value):
                        return None

                    if isinstance(value, float) and value.is_integer():
                        return str(int(value))

                    return str(value).strip()
           
                
                # Preparar datos para comprobaciones y variables previas

                cache = {
                    'muestras': Muestra.objects.values_list('nom_lab',flat=True),
                    'volumenes_actuales':{
                        sample.nom_lab : sample.volumen_actual
                        for sample in Muestra.objects.all() if sample.volumen_actual != None 
                    },
                    'estados_actuales':{
                        sample.nom_lab 
                        for sample in Muestra.objects.all() if sample.estado_actual != 'Destruida' or sample.estado_actual != 'ENV' or sample.estado_actual != None or sample.estado_actual != 'DEST'
                    },
                    'centros_envio': agenda_envio.objects.values_list('centro','lugar')
                }

                total_filas = len(df)

                def gen_validar_envios():
                    filas_validas = []
                    errores = {}
                    nom_lab_excel = set()
                    numero_registros = 0

                    yield _progress_page_start('Validando envíos', total_filas)

                    # Recorrer las filas del excel para realizar la validación previa a la carga de datos
                    for idx, row in df.iterrows():
                        numero_registros += 1
                        fila = idx + 2 
                        errores[fila]={"bloqueantes":[],"advertencias":[]}
                        # Registrar en el excel el centro y lugar de envio seleccionados de la agenda de envios
                        row['centro_destino'] = centro_envio.centro
                        row['lugar_destino'] = centro_envio.lugar

                        datos = {
                            "nom_lab":norm(row['nom_lab']),
                            "volumen_enviado":norm_code(row['volumen_enviado']),
                            "unidad_volumen_enviado":norm(row['unidad_volumen_enviado']),
                            "concentracion_enviada":norm_code(row['concentracion_enviada']),
                            "unidad_concentracion_enviada":norm(row['unidad_concentracion_enviada'])
                        }
              
                        # Comprobar si los campos obligatorios han sido rellenados
                        obligatorios = ["nom_lab", "volumen_enviado", "unidad_volumen_enviado", "concentracion_enviada", "unidad_concentracion_enviada"]

                        for campo in obligatorios:
                            if not datos.get(campo):
                                errores[fila]["bloqueantes"].append(f"campo_obligatorio_vacio:{campo}")
                    
                         # Comprobar si los campos estan en el formato correcto
                        for campo in ['volumen_enviado', 'concentracion_enviada']:
                            if datos[campo] != None:
                                try:
                                    datos[campo]=float(datos[campo])
                                except (TypeError, ValueError):
                                    errores[fila]["bloqueantes"].append(f"formato_incorrecto:{campo}")
                    
                        # Comprobar que la muestra exista en la base de datos y no esté duplicada en el excel
                        nom_lab = datos["nom_lab"]
                        if nom_lab not in cache["muestras"]:
                            errores[fila]["bloqueantes"].append("muestra_inexistente")
                        if nom_lab in nom_lab_excel:
                            errores[fila]["bloqueantes"].append("muestra_duplicada_excel")
                        else:
                            nom_lab_excel.add(nom_lab)

                        # Comprobar que el volumen a enviar no sea mayor al actual
                        volumen_envio = datos["volumen_enviado"]
                        if nom_lab in cache["volumenes_actuales"]:
                            if volumen_envio > cache["volumenes_actuales"][nom_lab]:
                                errores[fila]["bloqueantes"].append("volumen_alto")

                        # Comprobar que el estado de la muestra sea 'Disponible' o 'Parcialmente enviada'
                        if nom_lab not in cache['estados_actuales']:
                            errores[fila]["bloqueantes"].append("estado_no_disponible")

                        # Rellenar con el centro y lugar de destino
                        datos['centro_destino'] = centro_envio.centro
                        datos['lugar_destino'] = centro_envio.lugar
                    
                        # Registrar filas validas
                        if not errores[fila]["bloqueantes"]:
                            datos['fila'] = fila
                            filas_validas.append(datos)

                        if _should_update(idx, total_filas):
                            yield _progress_update(idx + 1, total_filas)
                
                    request.session['filas_validas'] = filas_validas
                    request.session['errores'] = errores

                    mensajes = [{'level': 'info', 'text': f'El excel subido tiene {numero_registros} registros.'}]
                    numero_errores = sum(1 for f in errores if errores[f]['bloqueantes'])
                    if numero_errores > 0:
                        mensajes.append({'level': 'error', 'text': f'Pero contiene {numero_errores} filas con errores graves.'})
                    else:
                        mensajes.append({'level': 'success', 'text': 'Y no tiene errores en ningún campo.'})

                    request.session['confirmacion_pendiente'] = {
                        'template': 'confirmacion_upload_envio.html',
                        'context': {},
                        'mensajes': mensajes
                    }
                    request.session.save()
                    yield _progress_done(request.path + '?mostrar_confirmacion=1', 'Validación completada')

                return StreamingHttpResponse(gen_validar_envios(), content_type='text/html')
        # Si se solicita un excel de errores, este se rellena en base a los errores detectados durante la validación 
        elif 'excel_errores' in request.POST:
                # Leer los errores y el excel de la sesión
                errores = request.session.get('errores',[])
                excel_bytes = base64.b64decode(request.session.get('excel_file_base64'))
                excel_file = io.BytesIO(excel_bytes)
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                # Definir los estilos para pintar el excel
                FILL_ERROR_CELL = PatternFill("solid", fgColor="F5C2C7")  # rojo fuerte
                # Diccionario de mensajes
                MENSAJES_ERROR = {
                    "campo_obligatorio_vacio": "Campo obligatorio vacío",
                    "formato_incorrecto": "Formato incorrecto de un campo",
                    "fecha_invalida": "Fecha inválida (Formato correcto: DD-MM-AAAA)",
                    "muestra_inexistente": "La muestra no existe en la base de datos",
                    "muestra_duplicada_excel": "Muestra duplicada dentro del Excel",
                    "volumen_alto": "La muestra no tiene suficiente volumen para el envio",
                    "estado_no_disponible": "La muestra está enviada o destruida, o no tiene un estado definido",
                    "caracter_invalido_semicolon": "El carácter ';' no está permitido en este campo"
                }
                # Diccionario de columnas del excel
                columnas_excel = {}
                rename_columns = {
                    'Muestra':'nom_lab',
                    'Volumen enviado':'volumen_enviado', 
                    'Volumen actual': 'volumen_actual',
                    'Concentración actual':'concentracion_actual',
                    'Concentración enviada':'concentracion_enviada',
                    'Unidad de volumen':'unidad_volumen_enviado',
                    'Unidad de concentración':'unidad_concentracion_enviada',
                    'Centro de destino':'centro_destino',
                    'Lugar de destino':'lugar_destino'
                }
                for cell in ws[1]:
                    columnas_excel[rename_columns[cell.value]] = cell.column
                # Añadir la columna de errores
                col_errores = ws.max_column + 1
                ws.cell(row=1, column=col_errores, value="Errores")
                # Mapeo de errores sin campo a sus columnas específicas (envíos)
                error_campo_map_envio = {
                    "muestra_inexistente": "nom_lab",
                    "muestra_duplicada_excel": "nom_lab",
                    "volumen_alto": "volumen_enviado",
                    "estado_no_disponible": "nom_lab"
                }
                # Recorrer filas con errores 
                for fila, info in errores.items():
                    has_error = bool(info["bloqueantes"])
                    if not has_error:
                        continue

                    # Colorear celdas específicas y construir mensajes
                    mensajes = []
                    for err in info["bloqueantes"]:
                        if ":" in err:
                            tipo, campo = err.split(":")
                            msg = f"(Error) {MENSAJES_ERROR[tipo]}"
                            if msg not in mensajes:
                                mensajes.append(msg)
                            col = columnas_excel[campo]
                            celda = ws.cell(row=int(fila), column=col)
                            celda.fill = FILL_ERROR_CELL
                        else:
                            campo = error_campo_map_envio.get(err)
                            msg = f"(Error) {MENSAJES_ERROR[err]}"
                            if msg not in mensajes:
                                mensajes.append(msg)
                            if campo and campo in columnas_excel:
                                col_err = columnas_excel[campo]
                                ws.cell(row=int(fila), column=col_err).fill = FILL_ERROR_CELL
                    ws.cell(row=int(fila), column=col_errores, value="\n".join(mensajes))
                # Retornar el excel de errores 
                output = io.BytesIO()    
                wb.save(output)
                wb.close()
                response = HttpResponse(output.getvalue(),content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename="listado_errores_envio.xlsx"'
                return response         
    else:
        form = UploadExcel(request)
    template = loader.get_template('upload_excel_envios.html')     
    return HttpResponse(template.render({'form': form},request))


def historial_envios(request,muestra_id):
    # Vista para ver el historial de envíos de una muestra específica
    sample = Muestra.objects.get(id=muestra_id)
    envios = Envio.objects.filter(muestra=sample).order_by('-fecha_envio')
    # Calcular el volumen original y el volumen restante
    volumen_original = sample.volumen_actual + sum(envio.volumen_enviado for envio in envios)
    volumen_restante = sample.volumen_actual
    template = loader.get_template('historial_envios.html')
    context = {
        'muestra':sample,
        'envios':envios,
        'volumen_original':volumen_original,
        'volumen_restante':volumen_restante
    }
    return HttpResponse(template.render(context,request))

def agenda(request):
    # Vista para ver la agenda de envíos de muestras
    agenda_envios = agenda_envio.objects.all()
    template = loader.get_template('agenda.html')
    return HttpResponse(template.render({'agenda':agenda_envios},request))

def nuevo_centro(request):
    # Vista para añadir un nuevo centro a la agenda de envíos
    if request.method == 'POST':
        form = Centroform(request.POST)
        if form.is_valid():
            form.save()
            return redirect('agenda')
        else:
            messages.error(request, 'Hubo un error al añadir el centro.')
    else:
        form = Centroform()
    template = loader.get_template('nuevo_centro.html')
    return HttpResponse(template.render({'form':form},request))

def editar_centro(request, id_centro):
    # Vista para editar un centro existente en la agenda de envíos
    centro = agenda_envio.objects.get(id=id_centro)
    if request.method == 'POST':
        form = Centroform(request.POST, instance=centro)
        if form.is_valid():
            form.save()
            return redirect('agenda')
    else:
        form = Centroform(instance=centro)
    return render(request, 'editar_centro.html', {'form': form, 'centro': centro})

def eliminar_centro(request):
    # Vista para eliminar centros de la agenda de envíos
    if request.method=="POST":
        ids = request.POST.getlist('ids_centro')
        for centro_id in ids:
            centro = agenda_envio.objects.get(id=centro_id)
            centro.delete()
    return redirect('agenda')


@login_required
@permission_required('muestras.can_view_localizaciones_web')
def exportar_posiciones_libres(request):
    """Vista que genera y descarga un Excel con las posiciones libres (subposiciones vacías)
    de los congeladores seleccionados, mostrando la jerarquía completa."""
    congelador_ids = request.GET.getlist('congelador')
    if not congelador_ids:
        messages.error(request, 'No se ha seleccionado ningún congelador.')
        return redirect('localizaciones_todas')

    # Obtener las subposiciones vacías de los congeladores seleccionados con toda la jerarquía
    subposiciones_libres = (
        Subposicion.objects
        .filter(vacia=True, caja__rack__estante__congelador__id__in=congelador_ids)
        .select_related('caja__rack__estante__congelador')
        .order_by(
            'caja__rack__estante__congelador__congelador',
            'caja__rack__estante__numero',
            'caja__rack__posicion_rack_estante',
            'caja__rack__numero',
            'caja__posicion_caja_rack',
            'caja__numero',
            'numero',
        )
    )

    # Crear el libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Posiciones libres"

    # Cabeceras
    cabeceras = [
        'Congelador',
        'Estante',
        'Posición rack en estante',
        'Rack',
        'Posición caja en rack',
        'Caja',
        'Subposición disponible',
    ]
    header_fill = Font(bold=True)

    for col_num, cabecera in enumerate(cabeceras, 1):
        celda = ws.cell(row=1, column=col_num, value=cabecera)
        celda.font = header_fill

    # Rellenar filas
    for fila_num, sub in enumerate(subposiciones_libres, 2):
        caja = sub.caja
        rack = caja.rack
        estante = rack.estante
        congelador = estante.congelador
        ws.cell(row=fila_num, column=1, value=congelador.congelador)
        ws.cell(row=fila_num, column=2, value=estante.numero)
        ws.cell(row=fila_num, column=3, value=rack.posicion_rack_estante)
        ws.cell(row=fila_num, column=4, value=rack.numero)
        ws.cell(row=fila_num, column=5, value=caja.posicion_caja_rack)
        ws.cell(row=fila_num, column=6, value=caja.numero)
        ws.cell(row=fila_num, column=7, value=sub.numero)

    # Ajustar ancho de columnas
    for col_num, cabecera in enumerate(cabeceras, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(cabecera) + 4, 18)

    # Nombres de congeladores para el nombre del archivo
    congeladores_selec = Congelador.objects.filter(id__in=congelador_ids)
    nombres = '_'.join([c.congelador for c in congeladores_selec])
    nombre_archivo = f'posiciones_libres_{nombres}.xlsx'

    # Devolver la respuesta como descarga
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response